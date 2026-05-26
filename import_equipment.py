#!/usr/bin/env python3
"""
Script de importación de equipos desde Excel.

Uso:
  python3 import_equipment.py equipos.xlsx --dry-run   # solo muestra qué importaría
  python3 import_equipment.py equipos.xlsx --commit    # importa en la DB

Formato esperado del Excel (cualquier hoja, con encabezados en fila 1):
  Columnas requeridas:  codigo   nombre    planta
  Columnas opcionales:  categoria   tipo_trabajo   marca   modelo   notas

  - codigo:         MC1, CG5, A7  (se normaliza: strip + uppercase)
  - planta:         MTR1 o MTR2
  - tipo_trabajo:   261-266

Tipos de trabajo conocidos:
  261 - Equipos móviles
  262 - Máquina de coser, mezcladora, grampas
  263 - Tolvas
  264 - Cintas
  265 - Sistema
  266 - Mantenimiento general de planta
"""

import sys
import os
import argparse

from dotenv import load_dotenv
load_dotenv()

try:
    import openpyxl
except ImportError:
    print("✗ Falta openpyxl. Instalalo con: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from app.database import engine, Base, SessionLocal
from app import models

# Alias de columnas aceptados (minúsculas sin tildes)
COL_ALIASES = {
    "codigo":         ["codigo", "code", "cod", "id_equipo", "equipo"],
    "nombre":         ["nombre", "name", "descripcion", "descripción", "description", "equipo_nombre"],
    "planta":         ["planta", "plant", "sede"],
    "categoria":      ["categoria", "categoría", "category", "tipo", "type"],
    "tipo_trabajo":   ["tipo_trabajo", "work_type", "trabajo", "codigo_trabajo", "cod_trabajo"],
    "marca":          ["marca", "brand"],
    "modelo":         ["modelo", "model", "model_name"],
    "notas":          ["notas", "notes", "observaciones"],
}

WORK_TYPE_CODES = {261, 262, 263, 264, 265, 266}


def normalize_header(h: str) -> str:
    if h is None:
        return ""
    return (h.strip()
              .lower()
              .replace("á", "a").replace("é", "e").replace("í", "i")
              .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))


def map_columns(headers: list) -> dict:
    """Devuelve {campo_canónico: índice_columna}."""
    norm = [normalize_header(h) for h in headers]
    mapping = {}
    for canon, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                mapping[canon] = norm.index(alias)
                break
    return mapping


def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_all = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            continue
        # Buscar fila de encabezados (primera fila con datos)
        header_row = None
        for i, row in enumerate(raw):
            if any(v is not None for v in row):
                header_row = i
                break
        if header_row is None:
            continue
        headers = [str(v) if v is not None else "" for v in raw[header_row]]
        mapping = map_columns(headers)
        if "codigo" not in mapping or "nombre" not in mapping:
            continue  # hoja sin columnas requeridas, saltar
        print(f"  ✓ Hoja '{sheet_name}': {len(raw) - header_row - 1} filas, columnas: {list(mapping.keys())}")
        for row in raw[header_row + 1:]:
            if all(v is None for v in row):
                continue
            entry = {}
            for canon, idx in mapping.items():
                entry[canon] = row[idx] if idx < len(row) else None
            rows_all.append(entry)
    wb.close()
    return rows_all


def parse_row(entry: dict, lineno: int) -> dict | None:
    """Parsea y valida una fila. Retorna None si hay error grave."""
    code_raw = entry.get("codigo")
    name_raw = entry.get("nombre")
    plant_raw = entry.get("planta")

    if code_raw is None or str(code_raw).strip() == "":
        print(f"  ⚠  Fila {lineno}: sin código — saltada")
        return None
    if name_raw is None or str(name_raw).strip() == "":
        print(f"  ⚠  Fila {lineno}: sin nombre — saltada")
        return None

    code = str(code_raw).strip().upper()
    name = str(name_raw).strip()
    plant = str(plant_raw).strip().upper() if plant_raw else "MTR1"
    if plant not in ("MTR1", "MTR2"):
        print(f"  ⚠  Fila {lineno} ({code}): planta '{plant}' no reconocida, se usa MTR1")
        plant = "MTR1"

    # tipo_trabajo
    wtc = None
    if entry.get("tipo_trabajo"):
        try:
            wtc = int(float(str(entry["tipo_trabajo"]).strip()))
            if wtc not in WORK_TYPE_CODES:
                print(f"  ⚠  Fila {lineno} ({code}): tipo_trabajo {wtc} fuera de rango 261-266")
                wtc = None
        except (ValueError, TypeError):
            pass

    return {
        "code": code,
        "name": name,
        "plant": plant,
        "category": str(entry["categoria"]).strip() if entry.get("categoria") else None,
        "work_type_code": wtc,
        "brand": str(entry["marca"]).strip() if entry.get("marca") else None,
        "model_name": str(entry["modelo"]).strip() if entry.get("modelo") else None,
        "notes": str(entry["notas"]).strip() if entry.get("notas") else None,
    }


def run(path: str, commit: bool):
    print(f"\n{'='*55}")
    print(f"  Importando equipos desde: {path}")
    print(f"  Modo: {'COMMIT' if commit else 'DRY-RUN'}")
    print(f"{'='*55}\n")

    # Crear tablas si no existen
    Base.metadata.create_all(bind=engine)

    raw_rows = read_excel(path)
    if not raw_rows:
        print("✗ No se encontraron filas con columnas requeridas (codigo, nombre).")
        sys.exit(1)

    print(f"\n  Total filas leídas: {len(raw_rows)}\n")

    parsed = []
    for i, entry in enumerate(raw_rows, start=2):
        result = parse_row(entry, i)
        if result:
            parsed.append(result)

    if not parsed:
        print("✗ Ninguna fila válida para importar.")
        sys.exit(1)

    db: Session = SessionLocal()
    try:
        existing_codes = {e.code for e in db.query(models.Equipment.code).all()}
        to_insert = []
        to_skip = []

        for item in parsed:
            if item["code"] in existing_codes:
                to_skip.append(item["code"])
            else:
                to_insert.append(item)

        print(f"  Equipos a importar:  {len(to_insert)}")
        print(f"  Duplicados saltados: {len(to_skip)}")
        if to_skip:
            print(f"  Duplicados: {', '.join(to_skip[:20])}")
        print()

        if not commit:
            print("  ── DRY-RUN: primeros 10 equipos a importar ──")
            for item in to_insert[:10]:
                print(f"    {item['code']:10} | {item['plant']:5} | {item['name'][:40]}")
            print("\n  Para importar de verdad, usá --commit")
            return

        # Commit
        for item in to_insert:
            eq = models.Equipment(**item)
            db.add(eq)

        db.commit()
        print(f"  ✓ {len(to_insert)} equipos importados correctamente.")

    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar equipos desde Excel")
    parser.add_argument("file", help="Ruta al archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true", default=False)
    parser.add_argument("--commit", action="store_true", default=False)
    args = parser.parse_args()

    if not args.commit and not args.dry_run:
        print("Especificá --dry-run o --commit")
        sys.exit(1)
    if not os.path.exists(args.file):
        print(f"✗ Archivo no encontrado: {args.file}")
        sys.exit(1)

    run(args.file, commit=args.commit)
