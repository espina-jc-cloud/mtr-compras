#!/usr/bin/env python3
"""
Import cargo summaries from 'Operativos barcos' Excel into operation_cargo_summaries.
CV is now explicit in the Excel — no inference needed.

Excel columns (0-indexed, Hoja 1):
  0  Nombre del Barco       → raw_ship_name, ship_name (normalized)
  1  Cliente                → client
  2  Producto               → product
  3  Fecha de Inicio        → start_date
  4  Fecha de Finalización  → end_date
  5  Cantidad de viajes     → trip_count
  6  Neto (kg)              → depot_kg  (canonical depot figure = balanza MTR)
  7  Origen (kg)            → ignored   (origin declared weight, not stored)
  8  Diferencias (kg)       → ignored
  9  Costado Vapor (kilos)  → cv_kg     (None if cell is empty)
  10 Total del Barco (kg)   → total_ship_kg

Idempotency key: (source_file, ship_name, client, product, start_date)

Usage:
    python3 import_cargo_summary.py --dry-run
    python3 import_cargo_summary.py --commit
    python3 import_cargo_summary.py --xlsx "other.xlsx" --source-file "operativos_2026.xlsx" --commit
"""
import argparse
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

DEFAULT_XLSX        = "/Users/juancruzespina/Downloads/Operativos barcos (1).xlsx"
DEFAULT_SOURCE_FILE = "operativos_barcos_2025_2026.xlsx"
ANOMALY_TOLERANCE_KG = 5_000    # max delta between total and depot+cv before warning
DATA_SHEET          = "Hoja 1"
EMPTY_STOP          = 3         # stop after N consecutive empty rows

# ── Product aliases: Excel name → Operation.product name in DB ────────────────
# Used only for matching, not for storing.
PRODUCT_MATCH_ALIASES: dict[str, str] = {
    "UREA":       "UREA GRANULADA",
    "TRIPLE":     "TRIPLE (STP)",
    "MAP (10-50)": "MAP ( 10-50)",
}

# Date window for matching operation by ship+date (days)
DATE_WINDOW = 10


# ── Normalization ─────────────────────────────────────────────────────────────

def norm_ship(s: str) -> str:
    """Normalize ship name: remove M/V prefix, uppercase, strip accents."""
    s = unicodedata.normalize("NFKD", str(s).strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    for prefix in ("M/V/", "M/V ", "MV ", "S/S "):
        if s.startswith(prefix):
            s = s[len(prefix):]
    return s.strip()


def to_dt(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d")
    except Exception:
        return None


def to_kg(val) -> int | None:
    """Convert Excel value (may be int, float, or None) to integer kg."""
    if val is None:
        return None
    try:
        v = float(val)
        return int(round(v)) if v >= 0 else None
    except (TypeError, ValueError):
        return None


# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_excel(xlsx_path: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("✗ openpyxl no instalado. Correr: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if DATA_SHEET in wb.sheetnames:
        ws = wb[DATA_SHEET]
    else:
        ws = wb.active
        print(f"  ⚠  Hoja '{DATA_SHEET}' no encontrada — usando hoja activa: {ws.title}")

    records = []
    empty_streak = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Stop on consecutive empty rows
        if not any(v is not None for v in row[:11]):
            empty_streak += 1
            if empty_streak >= EMPTY_STOP:
                break
            continue
        empty_streak = 0

        raw_ship = row[0]
        if raw_ship is None:
            continue
        raw_ship_str = str(raw_ship).strip()
        if not raw_ship_str:
            continue

        client     = str(row[1]).strip() if row[1] else None
        product    = str(row[2]).strip() if row[2] else "(sin producto)"
        start_date = to_dt(row[3])
        end_date   = to_dt(row[4])
        trip_count = to_kg(row[5])   # reuse int helper
        depot_kg   = to_kg(row[6])   # neto
        cv_raw     = row[9]          # Costado Vapor — may be None
        cv_kg      = to_kg(cv_raw) if cv_raw is not None else None
        total_kg   = to_kg(row[10])

        if depot_kg is None or total_kg is None:
            print(f"  ⚠  fila ignorada (depot o total vacío): {raw_ship_str} {product}")
            continue

        ship_normalized = norm_ship(raw_ship_str)

        # Anomaly: check total_ship ≈ depot + cv
        notes_parts = []
        if cv_kg is not None and cv_kg > 0:
            expected = depot_kg + cv_kg
            delta = abs(total_kg - expected)
            if delta > ANOMALY_TOLERANCE_KG:
                notes_parts.append(
                    f"ANOMALÍA: total_ship_kg ({total_kg:,}) no coincide con "
                    f"depósito + CV ({expected:,}), delta={delta:,} kg. Revisar Excel."
                )

        records.append({
            "raw_ship_name": raw_ship_str,
            "ship_name":     ship_normalized,
            "client":        client,
            "product":       product,
            "start_date":    start_date,
            "end_date":      end_date,
            "trip_count":    trip_count,
            "depot_kg":      depot_kg,
            "cv_kg":         cv_kg,
            "total_ship_kg": total_kg,
            "notes":         "; ".join(notes_parts) if notes_parts else None,
        })

    wb.close()
    return records


# ── Operation matching ────────────────────────────────────────────────────────

def match_operation(rec: dict, all_ops) -> tuple:
    """
    Try to match a record to an Operation row.
    Returns (operation_id, match_status).
    Matching criteria: ship_name (normalized) + product (with aliases) + date window.
    """
    target_ship  = rec["ship_name"]
    target_prod  = rec["product"]
    db_prod      = PRODUCT_MATCH_ALIASES.get(target_prod, target_prod)
    target_start = rec["start_date"]

    candidates = [
        op for op in all_ops
        if norm_ship(op.ship_name) == target_ship
    ]
    if not candidates:
        return None, "unmatched"

    # Filter by product
    prod_match = [op for op in candidates if op.product == db_prod or op.product == target_prod]
    if not prod_match:
        prod_match = candidates  # relax product constraint if no match

    # Filter by date window
    if target_start:
        date_match = [
            op for op in prod_match
            if op.start_date and abs((target_start - op.start_date).days) <= DATE_WINDOW
        ]
        if date_match:
            return date_match[0].id, "matched"

    # No date match — pick first candidate
    return prod_match[0].id, "matched"


# ── DB upsert ─────────────────────────────────────────────────────────────────

def run(xlsx_path: str, source_file: str, commit: bool):
    from app.database import SessionLocal
    from app import models

    print(f"Excel:       {xlsx_path}")
    print(f"Source file: {source_file}")
    print(f"Mode:        {'COMMIT' if commit else 'DRY-RUN'}")
    print()

    records = parse_excel(xlsx_path)
    print(f"Parsed {len(records)} rows from Excel")
    print()

    db = SessionLocal()
    try:
        all_ops = db.query(models.Operation).all()
        print(f"DB: {len(all_ops)} operations available for matching")
        print()

        stats = {"inserted": 0, "updated": 0, "skipped": 0, "warnings": 0}
        warnings = []

        for rec in records:
            op_id, match_status = match_operation(rec, all_ops)

            # Check for trip_count discrepancy vs matched Operation
            notes = rec["notes"] or ""
            if op_id:
                op = next((o for o in all_ops if o.id == op_id), None)
                if op and rec["trip_count"] and op.actual_trips:
                    diff = abs(rec["trip_count"] - op.actual_trips)
                    if diff > 0:
                        note = f"trip_count discrepancy: Excel={rec['trip_count']} DB={op.actual_trips}"
                        notes = (notes + "; " + note).strip("; ")
                        warnings.append(f"  ⚠  {rec['ship_name']} {rec['product']}: {note}")

            # Idempotency lookup
            existing = db.query(models.OperationCargoSummary).filter(
                models.OperationCargoSummary.source_file == source_file,
                models.OperationCargoSummary.ship_name   == rec["ship_name"],
                models.OperationCargoSummary.client      == rec["client"],
                models.OperationCargoSummary.product     == rec["product"],
                models.OperationCargoSummary.start_date  == rec["start_date"],
            ).first()

            if existing:
                # Check if any value changed
                changed = (
                    existing.depot_kg      != rec["depot_kg"]
                    or existing.cv_kg      != rec["cv_kg"]
                    or existing.total_ship_kg != rec["total_ship_kg"]
                    or existing.trip_count != rec["trip_count"]
                    or existing.operation_id != op_id
                )
                if changed:
                    existing.depot_kg      = rec["depot_kg"]
                    existing.cv_kg         = rec["cv_kg"]
                    existing.total_ship_kg = rec["total_ship_kg"]
                    existing.trip_count    = rec["trip_count"]
                    existing.end_date      = rec["end_date"]
                    existing.operation_id  = op_id
                    existing.match_status  = match_status
                    existing.notes         = notes or None
                    stats["updated"] += 1
                    action = "UPDATE"
                else:
                    stats["skipped"] += 1
                    action = "SKIP  "
            else:
                db.add(models.OperationCargoSummary(
                    source_file   = source_file,
                    raw_ship_name = rec["raw_ship_name"],
                    ship_name     = rec["ship_name"],
                    client        = rec["client"],
                    product       = rec["product"],
                    start_date    = rec["start_date"],
                    end_date      = rec["end_date"],
                    trip_count    = rec["trip_count"],
                    depot_kg      = rec["depot_kg"],
                    cv_kg         = rec["cv_kg"],
                    total_ship_kg = rec["total_ship_kg"],
                    operation_id  = op_id,
                    match_status  = match_status,
                    notes         = notes or None,
                ))
                stats["inserted"] += 1
                action = "INSERT"

            cv_str = f"CV={rec['cv_kg']:>10,}" if rec["cv_kg"] else "CV=         -"
            warn_flag = " ⚠" if rec["notes"] else ""
            print(
                f"  {action} {rec['ship_name'][:25]:25s} | {rec['product']:10s} | "
                f"{rec['start_date'].strftime('%Y-%m-%d') if rec['start_date'] else 'sin fecha':10s} | "
                f"depot={rec['depot_kg']:>10,} | {cv_str} | total={rec['total_ship_kg']:>10,} | "
                f"{match_status:9s}{warn_flag}"
            )

        print()
        if warnings:
            stats["warnings"] = len(warnings)
            print("Warnings:")
            for w in warnings:
                print(w)
            print()

        print(f"Result: inserted={stats['inserted']} updated={stats['updated']} "
              f"skipped={stats['skipped']} warnings={stats['warnings']}")

        if commit:
            db.commit()
            print("✓ Committed to DB")
        else:
            db.rollback()
            print("✗ Dry-run — nothing written. Use --commit to save.")

    except Exception as e:
        db.rollback()
        import traceback
        print(f"✗ Error: {e}", file=sys.stderr)
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Import cargo summaries from Operativos barcos Excel")
    p.add_argument("--xlsx",        default=DEFAULT_XLSX,        help="Path to Excel file")
    p.add_argument("--source-file", default=DEFAULT_SOURCE_FILE, help="Idempotency namespace (file identifier)")
    g = p.add_mutually_exclusive_group()
    g.add_argument("--dry-run", action="store_true", default=True,  help="Show what would be imported (default)")
    g.add_argument("--commit",  action="store_true", default=False, help="Actually write to DB")
    args = p.parse_args()

    run(
        xlsx_path   = args.xlsx,
        source_file = args.source_file,
        commit      = args.commit,
    )
