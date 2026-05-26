#!/usr/bin/env python3
"""
Import integral del Excel "Control de Mantenimiento MTR.xlsx".

Importa:
  1. Equipos (hoja 'Equipos')
  2. Registros históricos de mantenimiento (hoja 'Mantenimiento')
  3. Órdenes de Trabajo externas (hoja 'Ordenes de Trabajo')

Uso:
  python3 import_maintenance_excel.py archivo.xlsx --dry-run
  python3 import_maintenance_excel.py archivo.xlsx --commit

Protecciones:
  - --dry-run nunca toca la DB.
  - Equipos: dedup por código (strip+uppercase).
  - Registros: dedup por (plant, work_date, work_type_code, equip_key, responsable, titulo).
  - OT: dedup por (nro_ot, detalle).
  - Filas problemáticas: se importan de forma segura y se listan en ADVERTENCIAS.
"""

import sys
import os
import argparse
import re
from datetime import datetime
from decimal import Decimal

from dotenv import load_dotenv
load_dotenv()

try:
    import openpyxl
except ImportError:
    print("✗ Falta openpyxl: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from app.database import engine, Base, SessionLocal
from app import models

# ── Constantes ───────────────────────────────────────────────────────────────

EXCEL_PATH_DEFAULT = '/Users/juancruzespina/Downloads/Control de Mantenimiento MTR.xlsx'

WORK_TYPE_CODES = {261, 262, 263, 264, 265, 266}

# Responsables que se consideran internos (case-insensitive)
INTERNAL_PERFORMERS = {"gustavo"}

# Responsables que se consideran contratistas
CONTRACTOR_PERFORMERS = {"raúl", "raul", "romero", "fernando"}

# ── Helpers ──────────────────────────────────────────────────────────────────

def _norm_code(s) -> str:
    """Normaliza código de equipo: strip + uppercase + colapsa espacios."""
    if not s:
        return ""
    return re.sub(r'\s+', '', str(s).strip().upper())


def _norm_yesno(s) -> bool | None:
    """SI/NO → True/False/None."""
    if s is None:
        return None
    v = str(s).strip().upper()
    if v in ("SI", "SÍ", "S", "YES", "Y", "1"):
        return True
    if v in ("NO", "N", "0"):
        return False
    return None


def _norm_plant(v) -> str:
    """MTR 1 / 1.0 → 'MTR1' etc."""
    if v is None:
        return "MTR1"
    s = str(v).strip()
    if '2' in s:
        return "MTR2"
    return "MTR1"


def _norm_wtc(v) -> int | None:
    """261.0 → 261, None → None."""
    try:
        n = int(float(str(v).strip()))
        return n if n in WORK_TYPE_CODES else None
    except (ValueError, TypeError):
        return None


def _norm_cost(v) -> Decimal | None:
    try:
        f = float(str(v).replace(',', '.'))
        return Decimal(str(round(f, 2))) if f > 0 else None
    except (ValueError, TypeError, AttributeError):
        return None


def _is_contractor(responsable: str) -> bool:
    return responsable.strip().lower() in CONTRACTOR_PERFORMERS


def _is_internal(responsable: str) -> bool:
    return responsable.strip().lower() in INTERNAL_PERFORMERS


# ── Leer Excel ───────────────────────────────────────────────────────────────

def read_equipos(wb) -> list[dict]:
    ws = wb['Equipos']
    rows = list(ws.iter_rows(values_only=True))
    equipos = []
    for i, r in enumerate(rows[1:], start=2):
        nombre = r[0]
        codigo = r[1]
        planta = r[2]
        if not nombre or not codigo or not planta:
            continue
        # Solo saltar si el nombre O el código parecen ser una cabecera, no por columnas adyacentes
        nombre_str = str(nombre).strip().lower()
        codigo_str = str(codigo).strip().lower()
        if nombre_str in ('equipo', 'equipo ', 'nombre') or codigo_str in ('codificación', 'codigo', 'cod'):
            continue
        code = _norm_code(codigo)
        if not code:
            continue
        equipos.append({
            "row": i,
            "code": code,
            "name": str(nombre).strip(),
            "plant": _norm_plant(planta),
        })
    return equipos


def read_mantenimiento(wb) -> list[dict]:
    ws = wb['Mantenimiento']
    rows = list(ws.iter_rows(values_only=True))
    records = []
    # Row 0 = headers
    for i, r in enumerate(rows[1:], start=2):
        fecha = r[0]
        mtr = r[1]
        cod_trabajo = r[2]
        cod_equipo = r[3]
        engrase = r[4]
        limpieza = r[5]
        trabajo = r[6]
        observaciones = r[7]
        responsable = r[8]
        # r[9] = firma (skip)
        costo = r[10] if len(r) > 10 else None
        notas_extra = r[11] if len(r) > 11 else None

        if not fecha or not trabajo:
            continue
        if not isinstance(fecha, datetime):
            continue

        # Responsable
        resp_str = str(responsable).strip() if responsable else "Desconocido"

        # cod_equipo puede tener múltiples separados por \n
        if cod_equipo:
            equip_codes = [_norm_code(c) for c in str(cod_equipo).split('\n') if c.strip()]
        else:
            equip_codes = [""]

        # description = trabajo + observaciones + notas_extra
        desc_parts = [str(trabajo).strip()]
        if observaciones:
            desc_parts.append(f"Obs: {str(observaciones).strip()}")
        if notas_extra:
            desc_parts.append(f"Nota: {str(notas_extra).strip()}")
        full_desc = "\n".join(p for p in desc_parts if p)

        for eq_code in equip_codes:
            records.append({
                "row": i,
                "work_date": fecha,
                "plant": _norm_plant(mtr),
                "work_type_code": _norm_wtc(cod_trabajo),
                "equip_code_raw": eq_code,        # normalized, for DB lookup
                "equip_code_display": str(cod_equipo).strip() if cod_equipo else "",
                "did_lubrication": _norm_yesno(engrase),
                "did_cleaning": _norm_yesno(limpieza),
                "title": str(trabajo).strip()[:200],
                "description": full_desc,
                "responsable": resp_str,
                "is_contractor": _is_contractor(resp_str),
                "total_cost": _norm_cost(costo),
            })
    return records


def read_ordenes_trabajo(wb) -> list[dict]:
    ws = wb['Ordenes de Trabajo']
    rows = list(ws.iter_rows(values_only=True))
    ots = []
    for i, r in enumerate(rows[1:], start=2):
        # Avoid duplicated rows (data repeated in cols 9+)
        fecha = r[0]
        proveedor = r[1]
        nro_ot = r[2]
        mtr = r[3]
        detalle = r[4]
        cantidad = r[5]
        importe = r[6]
        responsable = r[7]

        if not fecha or not detalle:
            continue
        if not isinstance(fecha, datetime):
            continue

        desc = str(detalle).strip()
        if cantidad:
            desc += f" (cant. {cantidad})"

        title = f"OT {nro_ot} — {str(detalle).strip()}"[:200] if nro_ot else str(detalle).strip()[:200]

        ots.append({
            "row": i,
            "work_date": fecha,
            "plant": _norm_plant(mtr),
            "nro_ot": str(nro_ot).strip() if nro_ot else None,
            "proveedor_nombre": str(proveedor).strip() if proveedor else "Desconocido",
            "title": title,
            "description": desc,
            "responsable": str(responsable).strip() if responsable else "",
            "total_cost": _norm_cost(importe),
        })
    return ots


# ── Import logic ─────────────────────────────────────────────────────────────

def run(path: str, commit: bool):
    print(f"\n{'='*60}")
    print(f"  MTR — Import integral de mantenimiento")
    print(f"  Archivo : {os.path.basename(path)}")
    print(f"  Modo    : {'COMMIT ← escribe en DB' if commit else 'DRY-RUN (no escribe nada)'}")
    print(f"{'='*60}\n")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    equipos_raw    = read_equipos(wb)
    mant_raw       = read_mantenimiento(wb)
    ot_raw         = read_ordenes_trabajo(wb)
    wb.close()

    print(f"  Leído del Excel:")
    print(f"    Equipos           : {len(equipos_raw)}")
    print(f"    Registros mant.   : {len(mant_raw)}")
    print(f"    Órdenes de trabajo: {len(ot_raw)}")
    print()

    Base.metadata.create_all(bind=engine)
    db: Session = SessionLocal()

    warnings = []  # (row, sheet, msg)

    # ── 1. Equipos ────────────────────────────────────────────────────────
    print("── 1. EQUIPOS ─────────────────────────────────────────────")
    existing_codes = {e.code for e in db.query(models.Equipment.code).all()}
    eq_to_insert = []
    eq_skipped = []

    for eq in equipos_raw:
        if eq["code"] in existing_codes:
            eq_skipped.append(eq["code"])
        else:
            eq_to_insert.append(eq)
            existing_codes.add(eq["code"])  # track for duplicates within file

    print(f"  A importar  : {len(eq_to_insert)}")
    print(f"  Ya existen  : {len(eq_skipped)}")
    if eq_skipped:
        print(f"  Duplicados  : {', '.join(sorted(eq_skipped)[:20])}")

    if not commit:
        print("\n  Vista previa (primeros 15):")
        for eq in eq_to_insert[:15]:
            print(f"    {eq['code']:<8} | {eq['plant']:<5} | {eq['name']}")

    # ── 2. Registros de Mantenimiento ─────────────────────────────────────
    print("\n── 2. REGISTROS DE MANTENIMIENTO ──────────────────────────")

    # Build code→equipment_id map (both existing + to-insert)
    code_to_eq_id: dict[str, int] = {}
    for e in db.query(models.Equipment).all():
        code_to_eq_id[e.code] = e.id

    # Existing records dedup key
    existing_mant_keys: set[tuple] = set()
    for rec in db.query(models.MaintenanceRecord).filter(
        models.MaintenanceRecord.deleted_at.is_(None)
    ).all():
        key = (
            rec.plant,
            rec.work_date.date() if rec.work_date else None,
            rec.work_type_code,
            rec.equipment_id or rec.equipment_text or "",
            rec.performed_by_text or "",
            (rec.title or "")[:80],
        )
        existing_mant_keys.add(key)

    # Find or create system admin user for entered_by_id
    admin_user = db.query(models.User).filter(
        models.User.role.in_(["superadmin", "admin"])
    ).order_by(models.User.id).first()
    if not admin_user:
        print("  ✗ ERROR: no hay usuario admin/superadmin en la DB. Ejecutá migrate.py primero.")
        db.close()
        sys.exit(1)
    entered_by_id = admin_user.id

    # Resolve Gustavo → user id if exists
    gustavo_user = db.query(models.User).filter(
        models.User.name.ilike("%gustavo%")
    ).first()

    mant_to_insert = []
    mant_skipped = []

    for rec in mant_raw:
        # Resolve equipment
        eq_id = None
        eq_text = None

        if rec["equip_code_raw"]:
            # Try exact match first
            if rec["equip_code_raw"] in code_to_eq_id:
                eq_id = code_to_eq_id[rec["equip_code_raw"]]
            else:
                # Try normalized (remove spaces)
                stripped = re.sub(r'\s+', '', rec["equip_code_raw"])
                if stripped in code_to_eq_id:
                    eq_id = code_to_eq_id[stripped]
                else:
                    eq_text = rec["equip_code_display"] or rec["equip_code_raw"]
                    warnings.append((rec["row"], "Mantenimiento",
                        f"Cod equipo '{rec['equip_code_raw']}' no encontrado en catálogo → guardado en equipment_text"))

        # Determine performed_by
        performed_by_id = None
        performed_by_text = rec["responsable"]
        if not rec["is_contractor"] and gustavo_user:
            if rec["responsable"].strip().lower() in INTERNAL_PERFORMERS:
                performed_by_id = gustavo_user.id
                performed_by_text = None

        # Usar equip_code_raw como discriminador (no eq_text, que puede ser idéntico para splits)
        dedup_key = (
            rec["plant"],
            rec["work_date"].date(),
            rec["work_type_code"],
            eq_id or rec["equip_code_raw"] or eq_text or "",
            performed_by_text or (gustavo_user.name if performed_by_id else ""),
            rec["title"][:80],
        )

        if dedup_key in existing_mant_keys:
            mant_skipped.append(f"row{rec['row']}")
            continue

        existing_mant_keys.add(dedup_key)
        mant_to_insert.append({
            "row": rec["row"],
            "plant": rec["plant"],
            "work_date": rec["work_date"],
            "work_type_code": rec["work_type_code"],
            "equipment_id": eq_id,
            "equipment_text": eq_text,
            "maintenance_type": "correctivo",
            "title": rec["title"],
            "description": rec["description"],
            "did_lubrication": rec["did_lubrication"],
            "did_cleaning": rec["did_cleaning"],
            "is_contractor": rec["is_contractor"],
            "performed_by_id": performed_by_id,
            "performed_by_text": performed_by_text,
            "total_cost": rec["total_cost"],
            "status": "cerrado",
            "entered_by_id": entered_by_id,
        })

    print(f"  A importar  : {len(mant_to_insert)}")
    print(f"  Duplicados  : {len(mant_skipped)}")

    if not commit:
        print("\n  Vista previa:")
        for r in mant_to_insert[:10]:
            eq_label = f"EQ#{r['equipment_id']}" if r['equipment_id'] else (r['equipment_text'] or '—')
            print(f"    {r['work_date'].strftime('%d/%m/%Y')} | {r['plant']} | {eq_label:<12} | {r['performed_by_text'] or 'Gustavo':<10} | {r['title'][:40]}")
        if len(mant_to_insert) > 10:
            print(f"    ... ({len(mant_to_insert)-10} más)")

    # ── 3. Órdenes de Trabajo ────────────────────────────────────────────
    print("\n── 3. ÓRDENES DE TRABAJO ──────────────────────────────────")

    # Dedup OT: por título truncado (contiene N° OT en el título)
    existing_ot_titles: set[str] = set()
    for rec in db.query(models.MaintenanceRecord).filter(
        models.MaintenanceRecord.deleted_at.is_(None),
        models.MaintenanceRecord.title.ilike("OT %"),
    ).all():
        existing_ot_titles.add((rec.title or "")[:60])

    # Find or create 'Gomeria 2000' supplier
    gomeria_supplier = db.query(models.Supplier).filter(
        models.Supplier.name.ilike("%gomeria 2000%")
    ).first()

    ot_to_insert = []
    ot_skipped = []

    for ot in ot_raw:
        title_key = ot["title"][:60]
        if title_key in existing_ot_titles:
            ot_skipped.append(title_key)
            continue
        existing_ot_titles.add(title_key)

        ot_to_insert.append({
            "row": ot["row"],
            "plant": ot["plant"],
            "work_date": ot["work_date"],
            "title": ot["title"],
            "description": ot["description"],
            "is_contractor": True,
            "contractor_company": ot["proveedor_nombre"],
            "performed_by_text": ot["proveedor_nombre"],
            "performed_by_id": None,
            "supervised_by_id": None,
            "total_cost": ot["total_cost"],
            "status": "cerrado",
            "maintenance_type": "correctivo",
            "entered_by_id": entered_by_id,
            "supplier_id": gomeria_supplier.id if gomeria_supplier else None,
        })

    print(f"  A importar  : {len(ot_to_insert)}")
    print(f"  Duplicados  : {len(ot_skipped)}")
    if not commit:
        for r in ot_to_insert:
            print(f"    {r['work_date'].strftime('%d/%m/%Y')} | {r['plant']} | {r['contractor_company']:<15} | {r['title'][:45]}")

    # ── Resumen advertencias ─────────────────────────────────────────────
    print("\n── ADVERTENCIAS ───────────────────────────────────────────")
    if warnings:
        for row, sheet, msg in warnings:
            print(f"  ⚠  {sheet} fila {row}: {msg}")
    else:
        print("  ✓ Sin advertencias.")

    # ── Totales ──────────────────────────────────────────────────────────
    print(f"\n── RESUMEN {'(DRY-RUN — nada se escribe)' if not commit else '(COMMIT)'} ─────────────────────")
    print(f"  Equipos a importar          : {len(eq_to_insert)}")
    print(f"  Registros mant. a importar  : {len(mant_to_insert)}")
    print(f"  Órdenes de trabajo a import.: {len(ot_to_insert)}")
    print(f"  Total registros nuevos       : {len(mant_to_insert) + len(ot_to_insert)}")
    print(f"  Advertencias                 : {len(warnings)}")

    if not commit:
        print(f"\n  Para importar de verdad, usá --commit")
        db.close()
        return

    # ── COMMIT ───────────────────────────────────────────────────────────
    print("\n── EJECUTANDO COMMIT ──────────────────────────────────────")

    # 1. Insert equipos
    eq_inserted = 0
    for eq in eq_to_insert:
        obj = models.Equipment(
            code=eq["code"],
            name=eq["name"],
            plant=eq["plant"],
            active=True,
        )
        db.add(obj)
        eq_inserted += 1
    db.flush()

    # Rebuild code_to_eq_id after inserting new equipment
    code_to_eq_id = {e.code: e.id for e in db.query(models.Equipment).all()}

    # Resolve equipment_ids for maintenance records
    for rec in mant_to_insert:
        if rec["equipment_id"] is None and rec["equipment_text"] is None:
            pass  # already resolved
        elif rec["equipment_id"] is None and rec["equipment_text"]:
            # Try again with the freshly inserted equipment
            candidate = _norm_code(rec["equipment_text"])
            if candidate in code_to_eq_id:
                rec["equipment_id"] = code_to_eq_id[candidate]
                rec["equipment_text"] = None

    # 2. Insert maintenance records
    mant_inserted = 0
    for rec in mant_to_insert:
        mr = models.MaintenanceRecord(
            plant=rec["plant"],
            title=rec["title"],
            description=rec["description"],
            work_date=rec["work_date"],
            work_type_code=rec["work_type_code"],
            equipment_id=rec["equipment_id"],
            equipment_text=rec["equipment_text"],
            maintenance_type=rec["maintenance_type"],
            did_lubrication=rec["did_lubrication"],
            did_cleaning=rec["did_cleaning"],
            is_contractor=rec["is_contractor"],
            performed_by_id=rec["performed_by_id"],
            performed_by_text=rec["performed_by_text"],
            total_cost=rec["total_cost"],
            status=rec["status"],
            entered_by_id=rec["entered_by_id"],
        )
        db.add(mr)
        db.flush()

        log = models.MaintenanceAuditLog(
            record_id=mr.id,
            user_id=rec["entered_by_id"],
            action="imported",
            comment=f"Importado desde Excel (fila {rec['row']})",
        )
        db.add(log)
        mant_inserted += 1

    # 3. Insert OT records
    ot_inserted = 0
    for rec in ot_to_insert:
        mr = models.MaintenanceRecord(
            plant=rec["plant"],
            title=rec["title"],
            description=rec["description"],
            work_date=rec["work_date"],
            is_contractor=True,
            contractor_company=rec["contractor_company"],
            performed_by_text=rec["performed_by_text"],
            total_cost=rec["total_cost"],
            status=rec["status"],
            maintenance_type=rec["maintenance_type"],
            entered_by_id=rec["entered_by_id"],
            supplier_id=rec["supplier_id"],
        )
        db.add(mr)
        db.flush()

        log = models.MaintenanceAuditLog(
            record_id=mr.id,
            user_id=rec["entered_by_id"],
            action="imported",
            comment=f"Importado desde Excel OT (fila {rec['row']})",
        )
        db.add(log)
        ot_inserted += 1

    db.commit()
    print(f"  ✓ {eq_inserted} equipos insertados")
    print(f"  ✓ {mant_inserted} registros de mantenimiento insertados")
    print(f"  ✓ {ot_inserted} órdenes de trabajo insertadas")
    print(f"\n  Total: {eq_inserted + mant_inserted + ot_inserted} registros en DB.")

    if warnings:
        print(f"\n  ⚠  {len(warnings)} advertencias — revisá el listado de arriba.")

    # Verify
    final_eq = db.query(models.Equipment).count()
    final_mr = db.query(models.MaintenanceRecord).filter(
        models.MaintenanceRecord.deleted_at.is_(None)
    ).count()
    print(f"\n  Verificación final:")
    print(f"    equipment:            {final_eq}")
    print(f"    maintenance_records:  {final_mr}")

    db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importar todo el Excel de mantenimiento MTR")
    parser.add_argument("file", nargs="?", default=EXCEL_PATH_DEFAULT,
                        help="Ruta al Excel (default: ~/Downloads/Control de Mantenimiento MTR.xlsx)")
    parser.add_argument("--dry-run", action="store_true", default=False)
    parser.add_argument("--commit", action="store_true", default=False)
    args = parser.parse_args()

    if not args.commit and not args.dry_run:
        print("Especificá --dry-run o --commit")
        sys.exit(1)
    if not os.path.exists(args.file):
        print(f"✗ Archivo no encontrado: {args.file}")
        sys.exit(1)

    run(args.file, commit=args.commit)
