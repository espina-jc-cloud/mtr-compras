"""
Router: Despachos (Cupos y Despachos operativos de camiones)

Endpoints:
  GET  /despachos                — listado + KPIs + filtros
  GET  /despachos/import         — form de importación
  POST /despachos/import         — procesar upload + preview
  POST /despachos/import/confirm — confirmar e insertar batch
  GET  /despachos/{rid}          — detalle / edición
  POST /despachos/{rid}/status   — cambiar estado
  POST /despachos/{rid}/edit     — guardar cambios
  POST /despachos/{rid}/reprogram — reprogramar a otra fecha
"""

import io
import hashlib
import uuid
import os
from datetime import date, datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.deps import require_role
from app.permissions import require_perm
from app.models_cupos import (
    CupoDespacho, ImportBatch,
    DESPACHO_ESTADOS, DESPACHO_ESTADO_LABELS, ESTADO_CSS,
)
from app.templates import templates

router = APIRouter(prefix="/despachos", tags=["despachos"])

# Acceso al módulo Despachos → permiso "operaciones.despachos".
_guard = require_perm("operaciones.despachos")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# Helpers de importación
# ══════════════════════════════════════════════════════════════════════════════

def _detect_source(wb: openpyxl.Workbook) -> str:
    """
    Detecta si el archivo es de Nutrien, CNA o Plantilla MTR (combo).

    'combo' = archivo con hoja NUTRIEN + hoja Despachos (la Plantilla_Despachos_MTR.xlsx).
    """
    names = [s.lower() for s in wb.sheetnames]

    has_nutrien_sheet = any(
        k in n for n in names
        for k in ("nutrien", "ganel", "embolsado", "ramallo")
    ) or "base" in names

    has_cna_sheet = "despachos" in names

    # Plantilla MTR: tiene ambas hojas → "combo"
    if has_nutrien_sheet and has_cna_sheet:
        return "combo"

    if has_cna_sheet:
        return "cna"
    if has_nutrien_sheet:
        return "nutrien"

    # Fallback: si la primera hoja tiene columna IN/OUT → CNA
    ws = wb.active
    row1 = [str(ws.cell(1, c).value or "").strip().upper() for c in range(1, 10)]
    if "IN/OUT" in row1:
        return "cna"
    return "nutrien"


def _find_header_row(ws, keyword_col: str, max_scan: int = 20) -> Optional[int]:
    """
    Encuentra la primera fila que contenga keyword_col como encabezado.
    Retorna None si no encuentra.
    """
    keyword = keyword_col.upper().strip()
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value or "").upper().strip()
            if val == keyword:
                return r
    return None


def _normalize_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    # Ignorar filas de ejemplo de la plantilla MTR
    if s.startswith("←"):
        return None
    return s if s and s.upper() not in ("NONE", "N/A", "NULL") else None


def _normalize_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, (datetime,)):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _normalize_num(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _row_hash(*parts) -> str:
    joined = "|".join(str(p or "").strip().upper() for p in parts)
    return hashlib.sha256(joined.encode()).hexdigest()[:32]


# ─── Parser Nutrien ───────────────────────────────────────────────────────────

def _sheet_max_date(ws, fc: int, hrow: int) -> Optional[date]:
    """Retorna la fecha más reciente de una hoja (columna fc, a partir de hrow+1)."""
    latest = None
    for r in range(hrow + 1, min(ws.max_row + 1, hrow + 500)):
        v = _normalize_date(ws.cell(r, fc).value)
        if v and (latest is None or v > latest):
            latest = v
    return latest


def _parse_nutrien(
    wb: openpyxl.Workbook,
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
) -> list[dict]:
    """
    Parsea los cupos de Nutrien desde la hoja operativa correcta.

    SELECCIÓN DE HOJA:
      Prioridad 1 — hojas tipo ganel/embolsado/ramallo/nutrien
        Tienen la estructura D1/D2/SM completa y siempre en MT.
        Se elige la primera que tenga datos recientes (< 180 días).
      Prioridad 2 — hoja BASE, solo si no hay hojas operativas
        Fallback limpio. Cuidado: algunas versiones del archivo
        tienen BASE con kg en lugar de MT → se detecta automáticamente.
      Se toma UNA sola hoja para evitar importación doble.

    FILAS QUE SE SALTAN:
      - D1 / D2  → ingredientes de mezcla (el SM es el total del camión)
      - Packaging → "EMBOLSADO DE 50 KG", "BOLSONES" sin cantidad real
      - Hojas con fecha máxima anterior a 180 días (AMSUL 2020, etc.)

    UNIDADES:
      Si la columna Cantidad tiene valores > 200, se asume que está en kg
      y se convierte a MT dividiendo por 1000.
    """
    from datetime import timedelta
    cutoff_stale = date.today() - timedelta(days=180)
    rows = []

    # ── Elegir una sola hoja ──────────────────────────────────────────────────
    # 1. Candidatas con estructura detallada (ganel, nutrien, embolsado, etc.)
    candidates = [
        n for n in wb.sheetnames
        if any(k in n.lower() for k in
               ("nutrien", "ganel", "embolsado", "ramallo", "amsul", "profertil"))
    ]
    # 2. Si no hay, usar BASE
    if not candidates and "BASE" in wb.sheetnames:
        candidates = ["BASE"]
    # 3. Si nada, usar la primera hoja
    if not candidates:
        candidates = [wb.sheetnames[0]]

    chosen_sheet = None
    for name in candidates:
        ws_t = wb[name]
        # Buscar fila de header
        ht = None
        for kw in ("ST / SD / OD", "Destinatario", "ST/SD", "Destinatario "):
            ht = _find_header_row(ws_t, kw, max_scan=15)
            if ht:
                break
        if not ht:
            continue
        # Buscar col de fecha y verificar si la hoja es reciente
        fc_t = None
        for c in range(1, ws_t.max_column + 1):
            if "FECHA" in str(ws_t.cell(ht, c).value or "").upper():
                fc_t = c
                break
        if not fc_t:
            chosen_sheet = name
            break  # sin fecha, aceptar
        mx = _sheet_max_date(ws_t, fc_t, ht)
        if mx and mx >= cutoff_stale:
            chosen_sheet = name
            break  # primera hoja reciente encontrada

    if not chosen_sheet:
        # Ninguna hoja reciente → igualmente tomar la primera con header
        for name in candidates:
            ws_t = wb[name]
            for kw in ("ST / SD / OD", "Destinatario", "ST/SD"):
                if _find_header_row(ws_t, kw, max_scan=15):
                    chosen_sheet = name
                    break
            if chosen_sheet:
                break
    if not chosen_sheet:
        return rows  # sin hoja parseable

    # ── Parsear la hoja elegida ──────────────────────────────────────────────
    for sheet_name in [chosen_sheet]:
        ws = wb[sheet_name]

        # Buscar fila de cabecera (contiene "ST" o "Destinatario" o "Producto")
        hrow = None
        for keyword in ("ST / SD / OD", "Destinatario", "ST/SD", "Destinatario "):
            hrow = _find_header_row(ws, keyword, max_scan=15)
            if hrow:
                break
        if hrow is None:
            continue

        # Mapear columnas por nombre
        headers = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(hrow, c).value or "").strip()
            if val:
                headers[val] = c

        def col(name, aliases=()):
            names_to_try = [name] + list(aliases)
            for n in names_to_try:
                if n in headers:
                    return headers[n]
                # búsqueda parcial
                for h in headers:
                    if n.upper() in h.upper():
                        return headers[h]
            return None

        c_st       = col("ST / SD / OD", ("ST/SD", "S3", "O3"))
        c_dest     = col("Destinatario ", ("Destinatario", "DESTINATARIO"))
        c_prod     = col("Producto", ("PRODUCTO",))
        c_cant     = col("Cantidad 35(MT)", ("Cantidad (MT)", "CANTIDAD"))
        c_orig     = col("Origen", ("ORIGEN",))
        c_trans    = col("Transporte", ("TRANSPORTE",))
        c_fecha    = col("Fecha de carga", ("FECHA ENTREGA", "FECHA DE CARGA", "FECHA"))
        c_ac       = col("AC",)
        c_mezcla   = col("MEZCLA",)
        c_bolsa    = col("BOLSA",)
        c_npk      = col("N-P-K",)
        c_grado    = col("Grado",)
        c_pct      = col("%",)

        # Pre-leer todas las filas de datos para:
        # a) consultar fila siguiente (packaging lookahead)
        # b) detectar si las cantidades están en kg o MT
        data_rows = []
        for r in range(hrow + 1, ws.max_row + 1):
            sample = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 8))]
            if not any(v is not None for v in sample):
                continue
            data_rows.append(r)

        # ── Agrupación de camiones por BORDE GRUESO (regla real Nutrien) ──────────
        # Un camión multi-fila (mezcla/embolsado/consolidado) está encerrado por
        # bordes gruesos (medium/thick): empieza en una fila con top grueso y
        # termina en la fila con bottom grueso. Las filas SIN borde grueso son
        # camiones individuales (el granel monoproducto entero).
        def _thick_border(row_idx, edge):
            for c in range(1, ws.max_column + 1):
                s = getattr(ws.cell(row_idx, c).border, edge, None)
                if s and s.style in ("medium", "thick"):
                    return True
            return False

        row_grupo: dict = {}          # sheet_row → id de camión
        _gid = 0
        _cur = None
        _in_block = False
        for r in data_rows:
            top_thick = _thick_border(r, "top")
            bot_thick = _thick_border(r, "bottom")
            if _in_block:
                row_grupo[r] = _cur
                if bot_thick:
                    _in_block = False
            elif top_thick:
                _gid += 1
                _cur = _gid
                row_grupo[r] = _cur
                _in_block = not bot_thick   # si abre y cierra en la misma fila
            else:
                _gid += 1
                row_grupo[r] = _gid
        # ¿La hoja realmente usa bordes gruesos para agrupar? Si no hay ninguno,
        # cada fila es su propio grupo y el fallback por ST (import) sigue aplicando.
        _usa_bordes = any(
            _thick_border(r, "top") or _thick_border(r, "bottom") for r in data_rows
        )

        # Detectar unidad de cantidad: si el promedio de valores numéricos > 200 → kg
        cantidades = [
            float(ws.cell(r, c_cant).value)
            for r in data_rows
            if c_cant and ws.cell(r, c_cant).value is not None
            and isinstance(ws.cell(r, c_cant).value, (int, float))
            and float(ws.cell(r, c_cant).value) > 0
        ]
        cant_unit_is_kg = bool(cantidades) and (sum(cantidades) / len(cantidades)) > 200

        # Acumulador de ingredientes D1/D2 por ST_upper (para mezclas)
        d1d2_buffer: dict = {}  # st_upper → list of {producto, cant_mt}

        PACKAGING_KEYWORDS = ("EMBOLSADO", "BOLSONES", "BOLSON", "BOLSA DE", "EMBALAJE")

        for idx, r in enumerate(data_rows):
            fecha  = _normalize_date(ws.cell(r, c_fecha).value if c_fecha else None)
            st     = _normalize_str(ws.cell(r, c_st).value if c_st else None)
            prod   = _normalize_str(ws.cell(r, c_prod).value if c_prod else None)
            cant_raw = _normalize_num(ws.cell(r, c_cant).value if c_cant else None)
            cant = (cant_raw / 1000.0) if (cant_raw and cant_unit_is_kg) else cant_raw
            trans  = _normalize_str(ws.cell(r, c_trans).value if c_trans else None)
            mezcla = _normalize_str(ws.cell(r, c_mezcla).value if c_mezcla else None)
            bolsa  = _normalize_str(ws.cell(r, c_bolsa).value if c_bolsa else None)
            npk    = _normalize_str(ws.cell(r, c_npk).value if c_npk else None)
            grado  = _normalize_str(ws.cell(r, c_grado).value if c_grado else None)
            pct    = _normalize_num(ws.cell(r, c_pct).value if c_pct else None)

            # Normalizar "0" string (algunas hojas tienen ceros como placeholder)
            if prod in ("0", "0.0"):
                prod = None
            if st in ("0", "0.0"):
                st = None
            if not prod and not st:
                continue

            prod_upper = (prod or "").upper()
            st_upper   = (st   or "").upper()

            # ── Lógica de exclusión de sub-filas ─────────────────────────────
            # 1. D1 / D2 → acumular como componentes de la mezcla, luego skip
            if st and (st_upper.startswith("D1 ") or st_upper.startswith("D2 ")
                       or st_upper == "D1" or st_upper == "D2"):
                # Derivar la clave SM a partir del número de SM que comparten
                # Ej: "D1 26000194" → guardar bajo la clave del ST para asociar luego
                d1d2_buffer.setdefault(st_upper, []).append({
                    "producto": prod,
                    "cant_mt":  round(cant, 3) if cant else None,
                    "pct":      round(float(pct), 1) if pct else None,
                    "npk":      npk,
                })
                continue

            # 2. "SM XXXXX" (SM con número)
            if st_upper.startswith("SM ") and len(st_upper) > 3:
                if _usa_bordes:
                    # Formato consolidado (bordes): los COMPONENTES con cantidad
                    # real (FOSFATO, CLORURO, UREA…) son los productos del camión.
                    # Saltar solo los marcadores sin cantidad: "MEZCLADO DE
                    # FERTILIZANTES" y packaging (EMBOLSADO/BOLSONES).
                    if cant is None and ("MEZCLADO" in prod_upper
                                         or any(kw in prod_upper for kw in PACKAGING_KEYWORDS)):
                        continue
                    # con cantidad → se emite como producto normal más abajo
                elif "MEZCLADO" not in prod_upper:
                    # Formato plantilla (D1/D2/SM): solo conservar la fila MEZCLADO
                    continue

            # 3. Sub-filas de packaging sin cantidad real
            if cant is None and any(kw in prod_upper for kw in PACKAGING_KEYWORDS):
                continue

            # ── Filtro de fecha (Nutrien) ────────────────────────────────────
            if fecha:
                if date_from and fecha < date_from:
                    continue
                if date_to and fecha > date_to:
                    continue

            # ── Inferir presentación y bolsa_kg ─────────────────────────────
            presentacion = None
            bolsa_kg_val = None

            if bolsa and bolsa.upper() == "SI":
                # Lookahead: fila siguiente describe el envase
                if idx + 1 < len(data_rows):
                    next_r = data_rows[idx + 1]
                    next_prod = _normalize_str(ws.cell(next_r, c_prod).value if c_prod else None)
                    if next_prod:
                        np_upper = next_prod.upper()
                        if "BOLSONES" in np_upper or "BOLSON" in np_upper or "1000" in np_upper:
                            presentacion = "Bolsones 1000kg"
                            bolsa_kg_val = 1000
                        elif "50 KG" in np_upper or "50KG" in np_upper:
                            presentacion = "Bolsas 50kg"
                            bolsa_kg_val = 50
                        elif "25 KG" in np_upper or "25KG" in np_upper:
                            presentacion = "Bolsas 25kg"
                            bolsa_kg_val = 25
                        elif "EMBOLSADO" in np_upper:
                            presentacion = "Embolsado"
                if not presentacion:
                    # Intentar inferir del campo Grado si existe
                    if grado:
                        g = grado.upper()
                        if "50" in g:
                            presentacion = "Bolsas 50kg"; bolsa_kg_val = 50
                        elif "1000" in g or "BIG" in g:
                            presentacion = "Bolsones 1000kg"; bolsa_kg_val = 1000
                    if not presentacion:
                        presentacion = "Embolsado"
            elif mezcla and mezcla.upper() == "NO":
                presentacion = "Granel"

            # SM rows = mezcla granel
            is_sm = st and (st.upper() == "SM" or st_upper.startswith("SM "))
            if is_sm:
                presentacion = "Granel Mezcla"

            # ── Componentes de mezcla ────────────────────────────────────────
            # Para SM puro: los D1/D2 están en filas anteriores con distinto ST.
            # Buscamos en d1d2_buffer por prefijo de número si hay.
            componentes_json = None
            if is_sm:
                # Recolectar todos los D1/D2 acumulados que aún no se asignaron
                all_ingredientes = []
                for buf_key, buf_items in list(d1d2_buffer.items()):
                    all_ingredientes.extend(buf_items)
                if all_ingredientes:
                    import json as _json
                    componentes_json = _json.dumps(all_ingredientes, ensure_ascii=False)
                    d1d2_buffer.clear()
                # NPK de la fila SM tiene la fórmula completa
                if not npk and grado:
                    npk = grado

            rows.append({
                "source_type":    "nutrien",
                "document_type":  "cupo",
                "source_sheet":   sheet_name,
                "scheduled_date": fecha,
                "st_sd_od":       st,
                "external_ref":   st,
                "destinatario":   _normalize_str(ws.cell(r, c_dest).value if c_dest else None),
                "producto":       prod,
                "cantidad_mt":    cant,
                "origen":         _normalize_str(ws.cell(r, c_orig).value if c_orig else None),
                "transporte":     trans,
                "ac":             _normalize_str(ws.cell(r, c_ac).value if c_ac else None),
                "presentacion":   presentacion,
                "bolsa_kg":       bolsa_kg_val,
                "npk":            npk,
                "componentes_mezcla": componentes_json,
                "camion_grupo":   row_grupo.get(r) if _usa_bordes else None,
                "row_hash":       _row_hash("nutrien", fecha, st, prod, trans),
            })

    return rows


# ─── Parser CNA ───────────────────────────────────────────────────────────────

def _parse_cna(
    wb: openpyxl.Workbook,
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
) -> list[dict]:
    """
    Parsea la hoja 'Despachos' del Excel de CNA.
    Header en fila 1 con columnas fijas.
    """
    if "Despachos" not in wb.sheetnames:
        return []

    ws = wb["Despachos"]

    # Detectar fila de headers: fila 1 (CNA original) o fila 4 (Plantilla MTR)
    # La plantilla MTR tiene filas 1-3 como título/instrucciones
    def _find_header_row():
        for hrow in range(1, 8):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(hrow, c).value or "").strip().upper()
                if val in ("FECHA", "CLIENTE", "PRODUCTO", "KG. OC"):
                    return hrow
        return 1

    hrow = _find_header_row()
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(hrow, c).value or "").strip()
        if val:
            headers[val] = c

    def col(name):
        if name in headers:
            return headers[name]
        for h in headers:
            if name.upper() in h.upper():
                return headers[h]
        return None

    c_fecha   = col("Fecha")
    c_inout   = col("IN/OUT")
    c_client  = col("Cliente")
    c_cuit_c  = col("Cuit Cliente")
    c_np      = col("NP o FC")
    c_oc      = col("OC")
    c_prod    = col("Producto")
    c_kg_oc   = col("KG. OC")
    c_pres    = col("Presentacion")
    c_dest    = col("Destino")
    c_trans   = col("Transporte")
    c_cuit_t  = col("Cuit transporte")
    c_chofer  = col("Chofer")
    c_dni     = col("Dni chofer")
    c_chasis  = col("Pat. Chasis")
    c_acop    = col("Pat. Acoplado")
    c_obs     = col("Observaciones")
    c_neto    = col("Neto")
    c_remito  = col("Remito")

    rows = []
    for r in range(hrow + 1, ws.max_row + 1):
        fecha = _normalize_date(ws.cell(r, c_fecha).value if c_fecha else None)
        prod  = _normalize_str(ws.cell(r, c_prod).value if c_prod else None)
        if not fecha and not prod:
            continue
        # Ignorar filas de ejemplo de la plantilla MTR (prefijo ←)
        raw_fecha = str(ws.cell(r, c_fecha).value or "")
        if raw_fecha.strip().startswith("←"):
            continue

        # ── Filtro de fecha (CNA) ────────────────────────────────────────────
        if fecha:
            if date_from and fecha < date_from:
                continue
            if date_to and fecha > date_to:
                continue

        np_fc  = _normalize_str(ws.cell(r, c_np).value if c_np else None)
        trans  = _normalize_str(ws.cell(r, c_trans).value if c_trans else None)
        chasis = _normalize_str(ws.cell(r, c_chasis).value if c_chasis else None)

        # DNI: puede venir como float (Excel trata números grandes como float)
        dni_raw = ws.cell(r, c_dni).value if c_dni else None
        if isinstance(dni_raw, float):
            dni_raw = str(int(dni_raw))
        dni = _normalize_str(dni_raw)

        # CUIT cliente: similar problema
        cuit_c_raw = ws.cell(r, c_cuit_c).value if c_cuit_c else None
        if isinstance(cuit_c_raw, float):
            cuit_c_raw = str(int(cuit_c_raw))
        cuit_c = _normalize_str(cuit_c_raw)

        rows.append({
            "source_type":    "cna",
            "document_type":  "despacho",
            "source_sheet":   "Despachos",
            "scheduled_date": fecha,
            "actual_date":    fecha,   # CNA registra hechos consumados
            "in_out":         _normalize_str(ws.cell(r, c_inout).value if c_inout else None),
            "cliente":        _normalize_str(ws.cell(r, c_client).value if c_client else None),
            "cuit_cliente":   cuit_c,
            "external_ref":   np_fc,
            "order_number":   _normalize_str(ws.cell(r, c_oc).value if c_oc else None),
            "producto":       prod,
            "kg_oc":          _normalize_num(ws.cell(r, c_kg_oc).value if c_kg_oc else None),
            "presentacion":   _normalize_str(ws.cell(r, c_pres).value if c_pres else None),
            "destino":        _normalize_str(ws.cell(r, c_dest).value if c_dest else None),
            "transporte":     trans,
            "cuit_transporte": _normalize_str(ws.cell(r, c_cuit_t).value if c_cuit_t else None),
            "chofer":         _normalize_str(ws.cell(r, c_chofer).value if c_chofer else None),
            "dni_chofer":     dni,
            "patente_chasis":  chasis,
            "patente_acoplado": _normalize_str(ws.cell(r, c_acop).value if c_acop else None),
            "notes":          _normalize_str(ws.cell(r, c_obs).value if c_obs else None),
            "neto":           _normalize_num(ws.cell(r, c_neto).value if c_neto else None),
            "remito":         _normalize_str(ws.cell(r, c_remito).value if c_remito else None),
            "row_hash":       _row_hash("cna", fecha, np_fc, prod, trans, chasis),
            # CNA = registro consumado → estado cargado por defecto
            "status":         "cargado",
        })

    return rows


def _parse_excel(
    content:   bytes,
    filename:  str,
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
) -> tuple[str, list[dict], list[str]]:
    """
    Lee el Excel, detecta la fuente y parsea las filas.
    date_from / date_to: si se pasan, solo se incluyen filas dentro del rango.
    Retorna (source_type, rows, warnings).
    """
    warnings = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return "unknown", [], [f"Error al abrir el archivo: {e}"]

    source = _detect_source(wb)
    if source == "combo":
        # Plantilla MTR: parsear ambas hojas y combinar
        rows_n = _parse_nutrien(wb, date_from=date_from, date_to=date_to)
        rows_c = _parse_cna(wb, date_from=date_from, date_to=date_to)
        rows = rows_n + rows_c
        if not rows_n:
            warnings.append("Hoja NUTRIEN: sin datos (o fuera del rango de fecha).")
        if not rows_c:
            warnings.append("Hoja Despachos: sin datos (o fuera del rango de fecha).")
    elif source == "nutrien":
        rows = _parse_nutrien(wb, date_from=date_from, date_to=date_to)
    else:
        rows = _parse_cna(wb, date_from=date_from, date_to=date_to)

    if not rows:
        msg = "No se encontraron filas con datos válidos en el archivo."
        if date_from or date_to:
            msg += f" Revisá el rango de fechas ({date_from or '—'} → {date_to or '—'})."
        warnings.append(msg)

    return source, rows, warnings


# ══════════════════════════════════════════════════════════════════════════════
# Helpers de KPIs
# ══════════════════════════════════════════════════════════════════════════════

def _agrupar_camiones(registros: list) -> list:
    """
    Agrupa filas Nutrien por camion_grupo en un objeto camión con productos.
    Filas CNA se pasan como camiones de un solo producto.
    Devuelve lista de dicts listos para la vista.
    """
    from collections import defaultdict
    grupos: dict = defaultdict(list)
    orden_grupos: list = []

    for r in registros:
        if r.source_type == "nutrien" and r.camion_grupo is not None:
            key = ("nutrien", r.camion_grupo)
        else:
            key = ("single", r.id)
        if key not in grupos:
            orden_grupos.append(key)
        grupos[key].append(r)

    camiones = []
    for key in orden_grupos:
        filas = grupos[key]
        rep = filas[0]  # fila representativa para datos del camión
        total_t = sum(float(f.cantidad_mt or 0) + float(f.kg_oc or 0) / 1000 for f in filas)
        productos = [
            {
                "producto":     f.producto,
                "cantidad_mt":  f.cantidad_mt,
                "kg_oc":        f.kg_oc,
                "presentacion": f.presentacion,
                "bolsa_kg":     f.bolsa_kg,
                "npk":          f.npk,
            }
            for f in filas
        ]
        camiones.append({
            "id":            rep.id,
            "ids":           [f.id for f in filas],
            "source_type":   rep.source_type,
            "scheduled_date": rep.scheduled_date,
            "actual_date":   rep.actual_date,
            "cliente":       rep.cliente or rep.destinatario,
            "destinatario":  rep.destinatario,
            "transporte":    rep.transporte,
            "chofer":        rep.chofer,
            "patente_chasis": rep.patente_chasis,
            "patente_acoplado": rep.patente_acoplado,
            "status":        rep.status,
            "total_t":       round(total_t, 1),
            "productos":     productos,
            "multi":         len(filas) > 1,
            "camion_grupo":  rep.camion_grupo,
        })
    return camiones


def _kpis(registros: list) -> dict:
    from collections import Counter
    # contar camiones únicos (no filas)
    grupos_vistos: set = set()
    camiones_count = 0
    for r in registros:
        if r.source_type == "nutrien" and r.camion_grupo is not None:
            if r.camion_grupo not in grupos_vistos:
                grupos_vistos.add(r.camion_grupo)
                camiones_count += 1
        else:
            camiones_count += 1

    # para status usar el camión representativo (primer registro del grupo)
    reps: dict = {}
    for r in registros:
        key = r.camion_grupo if (r.source_type == "nutrien" and r.camion_grupo is not None) else r.id
        if key not in reps:
            reps[key] = r
    counter = Counter(r.status for r in reps.values())

    ton_prog = sum(float(r.cantidad_mt or 0) + float(r.kg_oc or 0) / 1000 for r in registros)
    ton_real = sum(float(r.neto or 0) / 1000 for r in registros if r.status == "cargado")
    return {
        "total":         camiones_count,
        "programado":    counter.get("programado", 0),
        "arribo":        counter.get("arribo", 0),
        "cargado":       counter.get("cargado", 0),
        "no_vino":       counter.get("no_vino", 0),
        "reprogramado":  counter.get("reprogramado", 0),
        "cancelado":     counter.get("cancelado", 0),
        "ton_programadas": round(ton_prog, 1),
        "ton_operadas":    round(ton_real, 1),
    }


# ══════════════════════════════════════════════════════════════════════════════
# Vista 1 — Listado principal
# ══════════════════════════════════════════════════════════════════════════════

@router.get("", response_class=HTMLResponse)
async def list_despachos(
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
    fecha_desde: str = "",
    fecha_hasta: str = "",
    cliente:     str = "",
    producto:    str = "",
    transporte:  str = "",
    status:      str = "",
    source_type: str = "",
    q:           str = "",
):
    query = db.query(CupoDespacho)

    if fecha_desde:
        try:
            query = query.filter(CupoDespacho.scheduled_date >= date.fromisoformat(fecha_desde))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            query = query.filter(CupoDespacho.scheduled_date <= date.fromisoformat(fecha_hasta))
        except ValueError:
            pass
    if cliente:
        query = query.filter(
            (CupoDespacho.cliente.ilike(f"%{cliente}%")) |
            (CupoDespacho.destinatario.ilike(f"%{cliente}%"))
        )
    if producto:
        query = query.filter(CupoDespacho.producto.ilike(f"%{producto}%"))
    if transporte:
        query = query.filter(CupoDespacho.transporte.ilike(f"%{transporte}%"))
    if status:
        query = query.filter(CupoDespacho.status == status)
    if source_type:
        query = query.filter(CupoDespacho.source_type == source_type)
    if q:
        query = query.filter(
            (CupoDespacho.cliente.ilike(f"%{q}%")) |
            (CupoDespacho.destinatario.ilike(f"%{q}%")) |
            (CupoDespacho.producto.ilike(f"%{q}%")) |
            (CupoDespacho.transporte.ilike(f"%{q}%")) |
            (CupoDespacho.patente_chasis.ilike(f"%{q}%")) |
            (CupoDespacho.chofer.ilike(f"%{q}%")) |
            (CupoDespacho.remito.ilike(f"%{q}%")) |
            (CupoDespacho.st_sd_od.ilike(f"%{q}%"))
        )

    registros = query.order_by(
        CupoDespacho.scheduled_date.desc(),
        CupoDespacho.id.desc()
    ).limit(500).all()

    kpis     = _kpis(registros)
    camiones = _agrupar_camiones(registros)

    return templates.TemplateResponse(
        request,
        "despachos/list.html",
        {
            "current_user":  current_user,
            "registros":     camiones, "shown": len(camiones), "truncated": len(camiones) >= 500,
            "kpis":          kpis,
            "estados":       DESPACHO_ESTADOS,
            "estado_labels": DESPACHO_ESTADO_LABELS,
            "estado_css":    ESTADO_CSS,
            # filtros activos
            "f_fecha_desde": fecha_desde,
            "f_fecha_hasta": fecha_hasta,
            "f_cliente":     cliente,
            "f_producto":    producto,
            "f_transporte":  transporte,
            "f_status":      status,
            "f_source":      source_type,
            "f_q":           q,
        },
    )


# ══════════════════════════════════════════════════════════════════════════════
# Vista 1b — Descargar plantilla
# ══════════════════════════════════════════════════════════════════════════════

# Ruta del archivo plantilla (generado con el script de creación)
_TEMPLATE_PATHS = [
    os.path.join(os.path.dirname(__file__), "..", "..", "Plantilla_Despachos_MTR.xlsx"),
    os.path.expanduser("~/Desktop/Plantilla_Despachos_MTR.xlsx"),
]

def _generate_template_bytes() -> bytes:
    """
    Genera el Excel plantilla en memoria.

    Diseño: columnas IDÉNTICAS a los archivos fuente reales para que el
    usuario pueda hacer copy/paste crudo sin reinterpretar nada.
    La inteligencia está en el parser, no en la persona que pega.

    Hoja NUTRIEN:  mismos headers que ganel_embolsado_ramallo (fila 10)
    Hoja Despachos: mismos headers que hoja Despachos de CNA (fila 1)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    GREEN_DARK = "1A5276"; GREEN_MED = "AED6F1"
    BLUE_DARK  = "154360"; BLUE_MED  = "A9CCE3"
    GRAY_INST  = "F2F3F4"; EJ_FG = "ABB2B9"; EJ_BG = "FDFEFE"

    def hdr(hex_bg):
        return dict(
            fill=PatternFill("solid", fgColor=hex_bg),
            font=Font(bold=True, color="FFFFFF", size=10),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
    def apply(cell, d):
        for a, v in d.items(): setattr(cell, a, v)
    def tb():
        s = Side(style="thin", color="D5D8DC")
        return Border(left=s, right=s, top=s, bottom=s)
    def ej_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=f"← {val}" if val is not None else "")
        c.fill = PatternFill("solid", fgColor=EJ_BG)
        c.font = Font(color=EJ_FG, italic=True, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = tb()

    # ── Hoja NUTRIEN ──────────────────────────────────────────────────────────
    # Headers IDÉNTICOS a ganel_embolsado_ramallo del Excel de Nutrien
    # El usuario selecciona desde la fila 11 de Nutrien y pega aquí desde la fila 5
    ws_n = wb.create_sheet("NUTRIEN")
    span_n = "A1:O1"
    ws_n.merge_cells(span_n); c = ws_n["A1"]
    c.value = ("PLANILLA OPERATIVA NUTRIEN — MTR  |  "
               "Copiá y pegá las filas de datos del Excel de Nutrien (fila 11 en adelante)")
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_n.row_dimensions[1].height = 20

    ws_n.merge_cells("A2:O2"); i2 = ws_n["A2"]
    i2.value = ("▶  Abrí el Excel de Nutrien → hoja ganel_embolsado_ramallo → "
                "seleccioná TODAS las filas de datos (fila 11 hasta el final, incluyendo "
                "filas D1/D2/SM/EMBOLSADO) → pegá aquí desde la fila 5. "
                "El importador detecta automáticamente qué es un camión real y qué es sub-fila.")
    i2.fill = PatternFill("solid", fgColor=GREEN_MED)
    i2.font = Font(italic=True, color=GREEN_DARK, size=9)
    i2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_n.row_dimensions[2].height = 36

    # Headers exactos del archivo real (fila 10 de ganel_embolsado_ramallo)
    nutrien_cols = [
        ("A", "Nº",               5),
        ("B", "ST / SD / OD",    16),
        ("C", "Destinatario ",   24),   # nota: espacio al final = igual que en el Excel real
        ("D", "Producto",        22),
        ("E", "Cantidad 35(MT)", 14),
        ("F", "%",                7),
        ("G", "MEZCLA",           8),
        ("H", "BOLSA",            8),
        ("I", "N-P-K",           22),
        ("J", "IA",               8),
        ("K", "Grado",           10),
        ("L", "Origen",          18),
        ("M", "Transporte",      18),
        ("N", "Fecha de carga",  13),
        ("O", "AC",              20),
    ]
    ws_n.merge_cells("A3:O3"); label_n = ws_n["A3"]
    label_n.value = "← copiá las filas de Nutrien a partir de aquí (fila 5)"
    label_n.fill = PatternFill("solid", fgColor="EAF4FC")
    label_n.font = Font(italic=True, color="5D6D7E", size=9)
    label_n.alignment = Alignment(horizontal="center", vertical="center")
    ws_n.row_dimensions[3].height = 14

    h_style_n = hdr(GREEN_DARK)
    for col, hdr_text, w in nutrien_cols:
        c = ws_n[f"{col}4"]
        c.value = hdr_text
        apply(c, h_style_n)
        c.border = tb()
        ws_n.column_dimensions[col].width = w
    ws_n.row_dimensions[4].height = 28

    # Filas de ejemplo (prefijo ← para que el parser las ignore)
    ej_rows_n = [
        [1, "26002780", "O´HIGGINS", "SUPERFOSFATO TRIPLE", 35, None, None, None, None, None, None, "MTR RAMALLO", "TRANSRUTA", "02/06/2026", "O´HIGGINS"],
        [23, "D1 26000194", "LEDESMA SA", "CLORURO DE POTASIO", 14, 0.4, "SI", "NO", "24N-0P-24K", None, None, "MTR RAMALLO", "LOCAL", "01/06/2026", "PARANÁ"],
        [24, "D2 26000115", "LEDESMA SA", "UREA", 18.27, 0.522, None, None, None, None, None, "MTR RAMALLO", "LOCAL", "01/06/2026", "PARANÁ"],
        [25, "D2 26000115", "LEDESMA SA", "AZUFERTIL", 2.73, 0.078, None, None, None, None, None, "MTR RAMALLO", "LOCAL", "01/06/2026", "PARANÁ"],
        [26, "SM", "LEDESMA SA", "MEZCLADO DE FERTILIZANTE", 35, None, None, None, None, None, None, "MTR RAMALLO", "LOCAL", "01/06/2026", "PARANÁ"],
    ]
    for ri, row_vals in enumerate(ej_rows_n):
        for ci, v in enumerate(row_vals):
            ej_cell(ws_n, 5 + ri, ci + 1, v)
        ws_n.row_dimensions[5 + ri].height = 15

    for r in range(10, 3000):
        ws_n[f"N{r}"].number_format = 'DD/MM/YYYY'
        ws_n[f"E{r}"].number_format = '#,##0.##'

    ws_n.auto_filter.ref = "A4:O4"
    ws_n.freeze_panes = "A5"

    # ── Hoja CNA (Despachos) ──────────────────────────────────────────────────
    # Headers IDÉNTICOS a la hoja Despachos del Excel de CNA
    ws_c = wb.create_sheet("Despachos")
    span_c = "A1:S1"
    ws_c.merge_cells(span_c); c2 = ws_c["A1"]
    c2.value = ("PLANILLA OPERATIVA CNA — MTR  |  "
                "Copiá y pegá las filas de datos de la hoja Despachos del Excel de CNA")
    c2.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c2.font = Font(bold=True, color="FFFFFF", size=11)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws_c.row_dimensions[1].height = 20

    ws_c.merge_cells("A2:S2"); i3 = ws_c["A2"]
    i3.value = ("▶  Abrí el drive de CNA → buscá el archivo del día → hoja Despachos → "
                "seleccioná las filas del período operativo (ej: solo hoy) → pegá aquí desde la fila 5. "
                "Copiás todo tal cual, sin modificar nada.")
    i3.fill = PatternFill("solid", fgColor=BLUE_MED)
    i3.font = Font(italic=True, color=BLUE_DARK, size=9)
    i3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_c.row_dimensions[2].height = 36

    ws_c.merge_cells("A3:S3"); label_c = ws_c["A3"]
    label_c.value = "← copiá las filas de CNA a partir de aquí (fila 5)"
    label_c.fill = PatternFill("solid", fgColor="EAF3FB")
    label_c.font = Font(italic=True, color="5D6D7E", size=9)
    label_c.alignment = Alignment(horizontal="center", vertical="center")
    ws_c.row_dimensions[3].height = 14

    cna_cols = [
        ("A", "Fecha",           12),
        ("B", "IN/OUT",           8),
        ("C", "Cliente",         24),
        ("D", "Cuit Cliente",    16),
        ("E", "NP o FC",         14),
        ("F", "OC",              10),
        ("G", "Producto",        22),
        ("H", "KG. OC",          12),
        ("I", "Presentacion",    14),
        ("J", "Destino",         24),
        ("K", "Transporte",      18),
        ("L", "Cuit transporte", 16),
        ("M", "Chofer",          22),
        ("N", "Dni chofer",      12),
        ("O", "Pat. Chasis",     12),
        ("P", "Pat. Acoplado",   14),
        ("Q", "Observaciones",   26),
        ("R", "Neto",            12),
        ("S", "Remito",          18),
    ]
    h_style_c = hdr(BLUE_DARK)
    for col, hdr_text, w in cna_cols:
        c = ws_c[f"{col}4"]
        c.value = hdr_text
        apply(c, h_style_c)
        c.border = tb()
        ws_c.column_dimensions[col].width = w
    ws_c.row_dimensions[4].height = 28

    dv = DataValidation(type="list", formula1='"In,Out"', allow_blank=True)
    ws_c.add_data_validation(dv)
    dv.sqref = "B5:B3000"

    ej_c_vals = [
        "05/06/2026", "Out", "BERNER SA", "30-70793421-9", "546", None,
        "Urea Perlada Automotor", 22000, "Big Bag", "Arrecifes", "Propio",
        None, "Rivera Marcos Daniel", "32581817", "KBF 679", "JAP 634",
        "Bolsones duros", 22060, "00004-00000017"
    ]
    for ci, v in enumerate(ej_c_vals):
        ej_cell(ws_c, 5, ci + 1, v)
    ws_c.row_dimensions[5].height = 15

    for r in range(6, 3000):
        ws_c[f"A{r}"].number_format = 'DD/MM/YYYY'
        ws_c[f"H{r}"].number_format = '#,##0'
        ws_c[f"R{r}"].number_format = '#,##0'

    ws_c.auto_filter.ref = "A4:S4"
    ws_c.freeze_panes = "A5"

    # ── Hoja USO ──────────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("📋 USO")
    ws_r.sheet_properties.tabColor = "F39C12"
    ws_r.column_dimensions["A"].width = 4
    ws_r.column_dimensions["B"].width = 30
    ws_r.column_dimensions["C"].width = 58
    ws_r.merge_cells("A1:C1"); t = ws_r["A1"]
    t.value = "GUÍA DE USO — La inteligencia está en el importador, no en vos"
    t.fill = PatternFill("solid", fgColor="2C3E50")
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_r.row_dimensions[1].height = 26

    instrucciones = [
        ("★ REGLA DE ORO", "Pegá crudo. No filtrés, no limpies, no decidás qué fila vale. Eso lo hace el sistema."),
        ("", ""),
        ("🟢 NUTRIEN — PASOS", ""),
        ("Paso 1", "Abrís el Excel de Nutrien → hoja ganel_embolsado_ramallo"),
        ("Paso 2", "Seleccionás DESDE LA FILA 11 HASTA EL FINAL — todo crudo"),
        ("      ↳ incluye", "D1 · D2 · SM · EMBOLSADO · packaging · cualquier sub-fila"),
        ("      ↳ no excluyas", "Nada. El sistema descarta lo que no es un camión válido"),
        ("Paso 3", "Ctrl+C → pegás en la fila 5 de la hoja NUTRIEN de esta plantilla"),
        ("Listo.", "No escribas presentación, no borres filas, no reinterpretés nada"),
        ("", ""),
        ("🔵 CNA — PASOS", ""),
        ("Paso 1", "Abrís el archivo de CNA del drive → hoja Despachos"),
        ("Paso 2", "Seleccionás las filas del período operativo (ej: solo hoy)"),
        ("Paso 3", "Ctrl+C → pegás en la fila 5 de la hoja Despachos de esta plantilla"),
        ("Listo.", "No modifiques nada, no traduzcas columnas, pegá tal cual"),
        ("", ""),
        ("⚙️ QUÉ RESUELVE EL IMPORTADOR SOLO", ""),
        ("D1 / D2", "Ingredientes de mezcla → los ignora automáticamente"),
        ("SM puro", "Fila total del camión mezcla → la conserva"),
        ("SM con número", "Conserva solo la fila con MEZCLADO en el producto"),
        ("Packaging sin qty", "Sub-fila de embolsado/bolsones → la ignora e infiere presentación"),
        ("Presentación", "Infiere Bolsas 50kg / Bolsones 1000kg / Granel / Granel Mezcla del contexto"),
        ("kg vs MT", "Si el promedio de cantidades > 200 → divide por 1000 automáticamente"),
        ("Duplicados", "No importa dos veces el mismo registro (hash único por fila)"),
        ("", ""),
        ("⚠️ LO ÚNICO QUE HACÉS VOS", ""),
        ("Filtro de fecha", "Al importar, elegís la fecha del día para no traer histórico"),
        ("Guardar copia", "Renombrá antes de importar: Plantilla_YYYY-MM-DD.xlsx"),
    ]
    for i, (label, desc) in enumerate(instrucciones):
        row = i + 2
        ws_r.row_dimensions[row].height = 18
        b = ws_r.cell(row=row, column=2, value=label)
        c = ws_r.cell(row=row, column=3, value=desc)
        is_regla = label.startswith("★")
        is_section = any(label.startswith(x) for x in ("🟢", "🔵", "⚙️", "⚠️"))
        if is_regla:
            for el in (b, c):
                el.fill = PatternFill("solid", fgColor="E74C3C")
                el.font = Font(bold=True, color="FFFFFF", size=11)
            ws_r.row_dimensions[row].height = 22
        elif is_section:
            for el in (b, c):
                el.fill = PatternFill("solid", fgColor="2C3E50")
                el.font = Font(bold=True, color="FFFFFF", size=10)
        elif label:
            bg = "F2F3F4" if i % 2 == 0 else "FDFEFE"
            b.fill = PatternFill("solid", fgColor=bg); b.font = Font(bold=True, color="2C3E50", size=10)
            c.fill = PatternFill("solid", fgColor=bg); c.font = Font(color="34495E", size=10)
        for el in (b, c):
            el.alignment = Alignment(vertical="center", wrap_text=(el == c))
    ws_r.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template")
async def download_template(
    current_user=Depends(_guard),
):
    """Descarga la Plantilla_Despachos_MTR.xlsx directamente desde el sistema."""
    content = _generate_template_bytes()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Plantilla_Despachos_MTR.xlsx"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Vista 2 — Importador: upload
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/import", response_class=HTMLResponse)
async def import_form(
    request: Request,
    current_user=Depends(_guard),
):
    return templates.TemplateResponse(
        request,
        "despachos/import.html",
        {"current_user": current_user, "preview": None, "error": None},
    )


@router.post("/import", response_class=HTMLResponse)
async def import_preview(
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
    file: UploadFile = File(...),
    source_override: str = Form("auto"),    # 'auto' | 'nutrien' | 'cna'
    import_fecha_desde: str = Form(""),     # filtro de fecha en importación
    import_fecha_hasta: str = Form(""),
):
    content = await file.read()

    # Parsear fechas de filtro
    df = None
    dt = None
    try:
        if import_fecha_desde:
            df = date.fromisoformat(import_fecha_desde)
    except ValueError:
        pass
    try:
        if import_fecha_hasta:
            dt = date.fromisoformat(import_fecha_hasta)
    except ValueError:
        pass

    source, rows, warnings = _parse_excel(content, file.filename, date_from=df, date_to=dt)

    if source_override != "auto":
        source = source_override

    if not rows:
        return templates.TemplateResponse(
            request,
            "despachos/import.html",
            {
                "current_user": current_user,
                "preview":      None,
                "error":        "No se encontraron datos en el archivo. " + " ".join(warnings),
            },
        )

    # Verificar duplicados existentes
    existing_hashes = {
        h for (h,) in db.query(CupoDespacho.row_hash).filter(
            CupoDespacho.row_hash.in_([r["row_hash"] for r in rows if r.get("row_hash")])
        ).all()
    }

    for r in rows:
        r["_duplicate"] = r.get("row_hash") in existing_hashes

    dup_count = sum(1 for r in rows if r["_duplicate"])
    new_count = len(rows) - dup_count

    # Guardar en sesión HTTP para el confirm
    import json
    # Serializar filas para la sesión (convertir date a str)
    def serial(obj):
        if isinstance(obj, date):
            return obj.isoformat()
        return obj

    rows_serializable = [
        {k: serial(v) for k, v in row.items()}
        for row in rows
    ]

    # Usar un campo hidden en el form para pasar el preview (max 200 filas para preview)
    preview_rows = rows[:200]

    return templates.TemplateResponse(
        request,
        "despachos/import.html",
        {
            "current_user":       current_user,
            "preview":            preview_rows,
            "all_rows_json":      json.dumps(rows_serializable),
            "source":             source,
            "filename":           file.filename,
            "total_rows":         len(rows),
            "dup_count":          dup_count,
            "new_count":          new_count,
            "warnings":           warnings,
            "error":              None,
            "estado_css":         ESTADO_CSS,
            "estado_labels":      DESPACHO_ESTADO_LABELS,
            "import_fecha_desde": import_fecha_desde,
            "import_fecha_hasta": import_fecha_hasta,
        },
    )


@router.post("/import/confirm")
async def import_confirm(
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
):
    form = await request.form()
    import json

    rows_json  = form.get("rows_json", "[]")
    source     = str(form.get("source", "unknown"))
    filename   = str(form.get("filename", ""))
    skip_dups  = form.get("skip_duplicates", "1") == "1"

    try:
        rows = json.loads(rows_json)
    except Exception:
        raise HTTPException(400, "Datos de importación inválidos.")

    if not rows:
        raise HTTPException(400, "No hay filas para importar.")

    # Crear batch
    batch_uuid = str(uuid.uuid4())
    batch = ImportBatch(
        batch_uuid=batch_uuid,
        source_type=source,
        filename=filename,
        sheet_name=rows[0].get("source_sheet", ""),
        imported_by=current_user.name,
    )
    db.add(batch)
    db.flush()

    # Verificar hashes existentes
    all_hashes = [r.get("row_hash") for r in rows if r.get("row_hash")]
    existing = {
        h for (h,) in db.query(CupoDespacho.row_hash).filter(
            CupoDespacho.row_hash.in_(all_hashes)
        ).all()
    } if all_hashes else set()

    # ── Agrupación de camiones (solo Nutrien) ─────────────────────────────────
    # Regla real: el camión se divide por BORDE GRUESO en la planilla (mezclas,
    # embolsados y consolidados), salvo el granel monoproducto entero que es un
    # camión por fila. Cuando el archivo trae esa info, el parser ya asignó
    # `camion_grupo` por borde → se respeta tal cual.
    #
    # Fallback (archivos pegados en la plantilla, sin bordes): agrupar por el
    # número de cupo ST / SD / OD.
    nutrien_rows = [(i, r) for i, r in enumerate(rows) if r.get("source_type") == "nutrien"]
    _tiene_bordes = any(r.get("camion_grupo") is not None for _, r in nutrien_rows)

    if not _tiene_bordes:
        _grupo_counter = 0
        st_to_grupo: dict = {}
        for i, r in nutrien_rows:
            st_key = (str(r.get("st_sd_od") or "").strip().upper()) or None
            if st_key is not None:
                if st_key not in st_to_grupo:
                    st_to_grupo[st_key] = _grupo_counter
                    _grupo_counter += 1
                r["camion_grupo"] = st_to_grupo[st_key]
            else:
                r["camion_grupo"] = _grupo_counter
                _grupo_counter += 1

    inserted = 0
    skipped  = 0

    for r in rows:
        h = r.get("row_hash")
        if skip_dups and h and h in existing:
            skipped += 1
            continue

        def to_date(v):
            if not v:
                return None
            if isinstance(v, str):
                try:
                    return date.fromisoformat(v)
                except ValueError:
                    return None
            return v

        cd = CupoDespacho(
            batch_id         = batch.id,
            source_type      = r.get("source_type", source),
            document_type    = r.get("document_type", "cupo"),
            row_hash         = h,
            scheduled_date   = to_date(r.get("scheduled_date")),
            actual_date      = to_date(r.get("actual_date")),
            st_sd_od         = r.get("st_sd_od"),
            external_ref     = r.get("external_ref"),
            order_number     = r.get("order_number"),
            remito           = r.get("remito"),
            cliente          = r.get("cliente"),
            cuit_cliente     = r.get("cuit_cliente"),
            destinatario     = r.get("destinatario"),
            destino          = r.get("destino"),
            ac               = r.get("ac"),
            producto         = r.get("producto"),
            cantidad_mt      = r.get("cantidad_mt"),
            kg_oc            = r.get("kg_oc"),
            neto             = r.get("neto"),
            presentacion     = r.get("presentacion"),
            bolsa_kg         = r.get("bolsa_kg"),
            npk              = r.get("npk"),
            componentes_mezcla = r.get("componentes_mezcla"),
            camion_grupo     = r.get("camion_grupo"),
            origen           = r.get("origen"),
            in_out           = r.get("in_out"),
            modo_transporte  = r.get("modo_transporte"),
            transporte       = r.get("transporte"),
            cuit_transporte  = r.get("cuit_transporte"),
            chofer           = r.get("chofer"),
            dni_chofer       = r.get("dni_chofer"),
            patente_chasis   = r.get("patente_chasis"),
            patente_acoplado = r.get("patente_acoplado"),
            notes            = r.get("notes"),
            status           = r.get("status", "programado"),
            imported_by      = current_user.name,
        )
        db.add(cd)
        inserted += 1

    batch.row_count = inserted
    db.commit()

    return RedirectResponse(
        url=f"/despachos?import_ok=1&inserted={inserted}&skipped={skipped}",
        status_code=303,
    )


# ══════════════════════════════════════════════════════════════════════════════
# Vista 3 — Detalle / Edición
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/{rid}", response_class=HTMLResponse)
async def detail(
    request: Request,
    rid: int,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
):
    reg = db.query(CupoDespacho).filter_by(id=rid).first()
    if not reg:
        raise HTTPException(404, "Registro no encontrado")

    return templates.TemplateResponse(
        request,
        "despachos/detail.html",
        {
            "current_user":  current_user,
            "reg":           reg,
            "estados":       DESPACHO_ESTADOS,
            "estado_labels": DESPACHO_ESTADO_LABELS,
            "estado_css":    ESTADO_CSS,
        },
    )


@router.post("/{rid}/status")
async def change_status(
    request: Request,
    rid: int,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
):
    reg = db.query(CupoDespacho).filter_by(id=rid).first()
    if not reg:
        raise HTTPException(404, "Registro no encontrado")

    form       = await request.form()
    new_status = str(form.get("status", "")).strip()
    notes      = str(form.get("notes", "")).strip()

    if new_status not in DESPACHO_ESTADO_LABELS:
        raise HTTPException(400, f"Estado inválido: {new_status}")

    reg.status = new_status
    if notes:
        reg.notes = notes
    reg.updated_at = datetime.utcnow()
    db.commit()

    # Redirect back
    back = str(form.get("back", "list"))
    if back == "detail":
        return RedirectResponse(url=f"/despachos/{rid}", status_code=303)
    return RedirectResponse(url="/despachos?ok=Estado+actualizado", status_code=303)


@router.post("/{rid}/edit")
async def edit_registro(
    request: Request,
    rid: int,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
):
    reg = db.query(CupoDespacho).filter_by(id=rid).first()
    if not reg:
        raise HTTPException(404, "Registro no encontrado")

    form = await request.form()

    def fget(name):
        v = form.get(name, "")
        return str(v).strip() or None

    def dget(name):
        v = fget(name)
        if not v:
            return None
        try:
            return date.fromisoformat(v)
        except ValueError:
            return None

    def nget(name):
        v = fget(name)
        if not v:
            return None
        try:
            return float(v)
        except ValueError:
            return None

    reg.actual_date      = dget("actual_date") or reg.actual_date
    reg.cliente          = fget("cliente")      or reg.cliente
    reg.destinatario     = fget("destinatario") or reg.destinatario
    reg.destino          = fget("destino")      or reg.destino
    reg.producto         = fget("producto")     or reg.producto
    reg.cantidad_mt      = nget("cantidad_mt")  or reg.cantidad_mt
    reg.neto             = nget("neto")         or reg.neto
    reg.presentacion     = fget("presentacion") or reg.presentacion
    reg.transporte       = fget("transporte")   or reg.transporte
    reg.chofer           = fget("chofer")       or reg.chofer
    reg.dni_chofer       = fget("dni_chofer")   or reg.dni_chofer
    reg.patente_chasis   = fget("patente_chasis")  or reg.patente_chasis
    reg.patente_acoplado = fget("patente_acoplado") or reg.patente_acoplado
    reg.remito           = fget("remito")       or reg.remito
    reg.notes            = fget("notes")        # puede ser None para borrar

    new_status = fget("status")
    if new_status and new_status in DESPACHO_ESTADO_LABELS:
        reg.status = new_status

    reg.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(url=f"/despachos/{rid}", status_code=303)


@router.post("/{rid}/reprogram")
async def reprogram(
    request: Request,
    rid: int,
    db: Session = Depends(get_db),
    current_user=Depends(_guard),
):
    reg = db.query(CupoDespacho).filter_by(id=rid).first()
    if not reg:
        raise HTTPException(404, "Registro no encontrado")

    form        = await request.form()
    nueva_fecha = str(form.get("nueva_fecha", "")).strip()
    notes       = str(form.get("notes", "")).strip()

    try:
        nf = date.fromisoformat(nueva_fecha)
    except ValueError:
        raise HTTPException(400, "Fecha inválida")

    # Actualizar registro original
    reg.status               = "reprogramado"
    reg.reprogrammed_to_date = nf
    if notes:
        reg.notes = (reg.notes or "") + f"\n[Reprogramado a {nf}: {notes}]".strip()
    reg.updated_at = datetime.utcnow()
    db.commit()

    return RedirectResponse(url=f"/despachos/{rid}", status_code=303)
