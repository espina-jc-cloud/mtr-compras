"""
Parser reutilizable de Operativos Finalizados (Excel de balanza).

Extraído de migrate._import_operations_history para poder importar operativos
desde la web (subiendo el Excel) además del auto-import histórico.

Estructura del Excel (0-indexed):
  col[0]  = Nombre del Barco (fila-header de operativo) / vacío en filas de viaje
  col[1]  = Cantidad de viajes declarada (en el header)
  col[2]  = código de viaje (int = viaje real; texto = subtotal → se ignora)
  col[3]  = Fecha entrada       col[4]  = Hora entrada
  col[5]  = Fecha salida        col[6]  = Hora salida
  col[7]  = Patente             col[8]  = Tara (kg)
  col[9]  = Bruto (kg)          col[10] = Neto (kg)
  col[11] = Origen (kg)         col[12] = Diferencias (kg)
  col[13] = Cliente             col[14] = Producto

Salida JSON-friendly (fechas como ISO str) para poder viajar en el campo hidden
del preview → confirm sin re-subir el archivo.
"""
from __future__ import annotations
import io
from datetime import datetime, timedelta
from collections import Counter


# ── Normalización (igual que el import histórico) ──────────────────────────────
SHIP_ALIASES = {
    "MV ARGENMAR MISTRAL":  "ARGENMAR MISTRAL",
    "M/V ARGENMAR MISTRAL": "ARGENMAR MISTRAL",
}
SPECIAL_NAMES = {"INGRESO SOP", "ZONA FRANCA", "DEVOLUCIÓN BUNGE"}
PRODUCT_FIX   = {"UREA GRANULA": "UREA GRANULADA"}


def _normalize_ship(name: str) -> str:
    s = (name or "").strip()
    return SHIP_ALIASES.get(s, s)


def _fix_product(val):
    if val is None:
        return None
    s = str(val).strip()
    return PRODUCT_FIX.get(s, s) if s else None


def _get_shift(t_str):
    if not t_str:
        return None
    try:
        h = int(str(t_str)[:2])
        if h < 6:  return 1
        if h < 12: return 2
        if h < 18: return 3
        return 4
    except Exception:
        return None


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d")
    except Exception:
        return None


def _to_int(val):
    try:
        return int(val) if val is not None else None
    except Exception:
        return None


def _calc_duration(entry_date, entry_time_str, exit_date, exit_time_str):
    if not all([entry_date, entry_time_str, exit_date, exit_time_str]):
        return None
    try:
        et = str(entry_time_str)[:8]
        xt = str(exit_time_str)[:8]
        entry_dt = datetime.combine(entry_date, datetime.strptime(et, "%H:%M:%S").time())
        exit_dt  = datetime.combine(exit_date,  datetime.strptime(xt, "%H:%M:%S").time())
        delta = exit_dt - entry_dt
        if delta.total_seconds() < 0:
            delta += timedelta(days=1)
        mins = round(delta.total_seconds() / 60, 2)
        return mins if 0 <= mins <= 1440 else None
    except Exception:
        return None


def _most_common(lst):
    clean = [x for x in lst if x]
    return Counter(clean).most_common(1)[0][0] if clean else None


def _iso(dt):
    return dt.isoformat() if dt else None


def _read_rows_xlsx(file_bytes: bytes):
    """Lee un .xlsx con openpyxl. Devuelve la lista de filas (tuplas de valores)."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl no está instalado en el servidor.")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("El archivo no es un Excel válido (.xlsx).")
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # saltar header
    wb.close()
    return rows


def _read_rows_xls(file_bytes: bytes):
    """Lee un .xls (formato viejo) con xlrd, normalizando celdas para que el
    parser las trate igual que las de openpyxl: fechas → datetime, horas → time."""
    try:
        import xlrd
    except ImportError:
        raise ValueError("El servidor no puede leer archivos .xls (falta xlrd).")
    try:
        book = xlrd.open_workbook(file_contents=file_bytes)
    except Exception:
        raise ValueError("El archivo no es un Excel válido (.xls).")

    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(1, sheet.nrows):  # saltar header
        cells = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            val = cell.value
            # ctype 3 = fecha/hora (número serial de Excel)
            if cell.ctype == 3:
                dt = xlrd.xldate.xldate_as_datetime(val, book.datemode)
                # Valor < 1 → solo hora (sin parte de fecha) → devolver time
                val = dt.time() if val < 1 else dt
            elif cell.ctype == 0:   # celda vacía
                val = None
            cells.append(val)
        rows.append(tuple(cells))
    return rows


def parse_operations_workbook(file_bytes: bytes):
    """Parsea el Excel de operativos (.xlsx o .xls). Devuelve (operations, warnings).

    `operations` es una lista de dicts JSON-friendly, uno por operativo, con sus
    viajes y los agregados ya calculados (toneladas, promedios, rango de fechas).
    Lanza ValueError si el archivo no se puede leer como planilla.
    """
    if not file_bytes:
        raise ValueError("El archivo está vacío.")

    # Detectar formato. Muchos sistemas (incluido el de balanza MTR) exportan
    # ".xls" que en realidad son HTML; los detectamos y parseamos aparte.
    head = file_bytes[:64].lstrip()[:16].lower()
    if head.startswith(b"<") and (b"<table" in file_bytes[:4096].lower()
                                  or b"<html" in head or b"<div" in head):
        operations_data, warnings = _read_operativos_html(file_bytes)
        return _finalize_operations(operations_data, warnings)

    # Excel binario: .xls = OLE2 (D0 CF 11 E0); .xlsx = ZIP (PK).
    magic = file_bytes[:4]
    if magic[:2] == b"PK":
        raw_rows = _read_rows_xlsx(file_bytes)
    elif magic == b"\xd0\xcf\x11\xe0":
        raw_rows = _read_rows_xls(file_bytes)
    else:
        try:
            raw_rows = _read_rows_xlsx(file_bytes)
        except ValueError:
            raw_rows = _read_rows_xls(file_bytes)

    warnings: list[str] = []
    operations_data: list[dict] = []
    current_op = None

    for row_idx, row in enumerate(raw_rows):
        def col(i):
            return row[i] if i < len(row) else None

        c0, c1, c2 = col(0), col(1), col(2)

        # Fila-header de operativo: col[0] con nombre y col[2] no numérico
        if c0 is not None and not isinstance(c2, (int, float)):
            raw_name = str(c0).strip()
            if raw_name:
                current_op = {
                    "raw_name":       raw_name,
                    "ship_name":      _normalize_ship(raw_name),
                    "operation_type": "special" if raw_name in SPECIAL_NAMES else "vessel",
                    "declared_trips": _to_int(c1),
                    "trips":          [],
                }
                operations_data.append(current_op)
            continue

        # Fila de viaje real: col[2] entero
        if isinstance(c2, (int, float)) and c2 == int(c2):
            if current_op is None:
                warnings.append(f"Fila {row_idx + 2}: viaje sin operativo padre — omitida.")
                continue

            entry_time_str = str(col(4))[:8] if col(4) is not None else None
            exit_time_str  = str(col(6))[:8] if col(6) is not None else None
            entry_date     = _to_date(col(3))
            exit_date      = _to_date(col(5))

            current_op["trips"].append({
                "trip_code":    int(c2),
                "entry_date":   _iso(entry_date),
                "entry_time":   entry_time_str,
                "exit_date":    _iso(exit_date),
                "exit_time":    exit_time_str,
                "plate":        str(col(7)).strip() if col(7) else None,
                "tara_kg":      _to_int(col(8)),
                "bruto_kg":     _to_int(col(9)),
                "neto_kg":      _to_int(col(10)),
                "origen_kg":    _to_int(col(11)),
                "diff_kg":      _to_int(col(12)),
                "shift_number": _get_shift(entry_time_str),
                "duration_min": _calc_duration(entry_date, entry_time_str, exit_date, exit_time_str),
                "client":       str(col(13)).strip() if col(13) else None,
                "product":      _fix_product(col(14)),
            })
        # else: subtotal u otra fila → se ignora en silencio

    return _finalize_operations(operations_data, warnings)


def _finalize_operations(operations_data: list[dict], warnings: list[str]):
    """Calcula los agregados por operativo y descarta los que no tienen viajes."""
    result: list[dict] = []
    for op in operations_data:
        trips = op["trips"]
        if not trips:
            continue

        neto_list   = [t["neto_kg"]   for t in trips if t["neto_kg"]   is not None]
        origen_list = [t["origen_kg"] for t in trips if t["origen_kg"] is not None]
        diff_list   = [t["diff_kg"]   for t in trips if t["diff_kg"]   is not None]
        dur_list    = [t["duration_min"] for t in trips
                       if t["duration_min"] is not None and 0 < t["duration_min"] <= 240]

        total_neto   = sum(neto_list)
        actual_trips = len(trips)

        entry_dates = [datetime.fromisoformat(t["entry_date"]) for t in trips if t["entry_date"]]
        start_date  = min(entry_dates) if entry_dates else None
        end_date    = max(entry_dates) if entry_dates else None
        op_hours    = max((end_date - start_date).total_seconds() / 3600, 1.0) if (start_date and end_date) else None

        op["client"]            = _most_common([t["client"]  for t in trips])
        op["product"]           = _most_common([t["product"] for t in trips])
        op["start_date"]        = _iso(start_date)
        op["end_date"]          = _iso(end_date)
        op["actual_trips"]      = actual_trips
        op["total_neto_kg"]     = total_neto
        op["total_origen_kg"]   = sum(origen_list)
        op["total_diff_kg"]     = sum(diff_list)
        op["avg_duration_min"]  = round(sum(dur_list) / len(dur_list), 2) if dur_list else None
        op["avg_tons_per_trip"] = round(total_neto / 1000 / actual_trips, 3) if actual_trips else None
        op["avg_tons_per_hour"] = round(total_neto / 1000 / op_hours, 3) if op_hours else None
        result.append(op)

    if not result:
        warnings.append("No se detectó ningún operativo con viajes en el archivo.")

    return result, warnings


def _ddmmyyyy(s):
    """Parsea fecha 'dd-mm-yyyy' o 'dd/mm/yyyy' → datetime (medianoche) o None."""
    if not s:
        return None
    s = str(s).strip().replace("/", "-")
    for fmt in ("%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _int_txt(s):
    """Convierte texto de celda HTML a int (tolera separadores) o None."""
    if s is None:
        return None
    s = str(s).strip().replace(".", "").replace(",", "")
    if not s or s in ("-", "."):
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return None


def _read_operativos_html(file_bytes: bytes):
    """Parsea el HTML exportado por el sistema de balanza MTR (.xls que es HTML).

    Estructura:
      - Una tabla de cabecera con 'Operativo: NN) NOMBRE BARCO (dd-mm-yy) ...'.
      - Una tabla de viajes con header [Nro, Fecha Ent, H. Ent, Fecha Sal, H. Sal,
        Pat. Cam, Tara, Bruto, Neto, Peso Orig, Diferencia, Cliente, Producto],
        filas 'Transporte: ...' (se ignoran) y filas de subtotal 'Viajes: ...'.
    """
    import re
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ValueError("El servidor no puede leer este formato (falta beautifulsoup4).")

    try:
        html = file_bytes.decode("latin-1")
    except Exception:
        html = file_bytes.decode("utf-8", errors="replace")

    soup = BeautifulSoup(html, "html.parser")
    warnings: list[str] = []

    # Nombre del operativo desde 'Operativo: NN) NOMBRE [(fecha)] Fecha Inicio: ...'.
    # El nombre termina en el '(' de la fecha o en la etiqueta 'Fecha' (lo que venga
    # primero). Sin este corte, el nombre se tragaría todo el texto del documento.
    full_text = soup.get_text(" ", strip=True)
    m = re.search(r"Operativo:\s*(?:\d+\)\s*)?(.+?)\s*(?:\(\d|\bFecha\b|\bCliente:|\bProducto:|\bTurno:)",
                  full_text)
    if not m:  # respaldo acotado (máx 60 chars) para no arrastrar la tabla entera
        m = re.search(r"Operativo:\s*(?:\d+\)\s*)?(.{1,60}?)\s{2,}", full_text)
    raw_name = (m.group(1).strip() if m else "Operativo importado")

    # Localizar la tabla de viajes (la que tiene el header con 'Neto').
    trip_rows = []
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        header = [c.get_text(" ", strip=True).lower() for c in (rows[0].find_all(["td", "th"]) if rows else [])]
        if any("neto" in h for h in header) and any("fecha" in h for h in header):
            trip_rows = rows[1:]  # saltar header
            break

    if not trip_rows:
        return [], ["No se encontró la tabla de viajes en el archivo."]

    op = {
        "raw_name":       raw_name,
        "ship_name":      _normalize_ship(raw_name),
        "operation_type": "special" if raw_name in SPECIAL_NAMES else "vessel",
        "declared_trips": None,
        "trips":          [],
    }

    for tr in trip_rows:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
        if len(cells) < 13:
            continue  # 'Transporte: ...' (1 celda) o subtotal (5 celdas) → ignorar
        code = _int_txt(cells[0])
        if code is None:
            continue  # fila que no es un viaje real

        entry_date = _ddmmyyyy(cells[1])
        exit_date  = _ddmmyyyy(cells[3])
        entry_time = cells[2].strip()[:8] or None
        exit_time  = cells[4].strip()[:8] or None

        op["trips"].append({
            "trip_code":    code,
            "entry_date":   _iso(entry_date),
            "entry_time":   entry_time,
            "exit_date":    _iso(exit_date),
            "exit_time":    exit_time,
            "plate":        cells[5].strip() or None,
            "tara_kg":      _int_txt(cells[6]),
            "bruto_kg":     _int_txt(cells[7]),
            "neto_kg":      _int_txt(cells[8]),
            "origen_kg":    _int_txt(cells[9]),
            "diff_kg":      _int_txt(cells[10]),
            "shift_number": _get_shift(entry_time),
            "duration_min": _calc_duration(entry_date, entry_time, exit_date, exit_time),
            "client":       cells[11].strip() or None,
            "product":      _fix_product(cells[12]),
        })

    return [op], warnings


def insert_operations(db, operations_data: list[dict], source_file: str):
    """Inserta los operativos parseados. Idempotente sobre trip_code (único global).

    Devuelve stats: operativos creados, viajes insertados, viajes duplicados saltados.
    """
    from app import models

    ops_created = trips_inserted = trips_dup = 0

    # trip_codes ya existentes (para saltar duplicados sin romper la transacción)
    incoming_codes = [t["trip_code"] for op in operations_data for t in op["trips"]]
    existing_codes = set()
    if incoming_codes:
        rows = (
            db.query(models.OperationTrip.trip_code)
            .filter(models.OperationTrip.trip_code.in_(incoming_codes))
            .all()
        )
        existing_codes = {r[0] for r in rows}

    def _dt(s):
        return datetime.fromisoformat(s) if s else None

    for op_data in operations_data:
        trips = op_data.get("trips", [])
        if not trips:
            continue

        # Viajes nuevos (los que no chocan con trip_codes ya existentes).
        new_trips = [t for t in trips if t["trip_code"] not in existing_codes]
        if not new_trips:
            # Todos los viajes ya estaban → no crear un operativo huérfano.
            trips_dup += len(trips)
            continue

        op = models.Operation(
            raw_name          = op_data["raw_name"],
            ship_name         = op_data["ship_name"],
            operation_type    = op_data.get("operation_type", "vessel"),
            client            = op_data.get("client"),
            product           = op_data.get("product"),
            start_date        = _dt(op_data.get("start_date")),
            end_date          = _dt(op_data.get("end_date")),
            declared_trips    = op_data.get("declared_trips"),
            actual_trips      = op_data.get("actual_trips", len(trips)),
            total_neto_kg     = op_data.get("total_neto_kg", 0),
            total_origen_kg   = op_data.get("total_origen_kg", 0),
            total_diff_kg     = op_data.get("total_diff_kg", 0),
            avg_duration_min  = op_data.get("avg_duration_min"),
            avg_tons_per_trip = op_data.get("avg_tons_per_trip"),
            avg_tons_per_hour = op_data.get("avg_tons_per_hour"),
            source_file       = source_file,
        )
        db.add(op)
        db.flush()
        ops_created += 1

        # Contar como duplicados los viajes de este op que sí existían.
        trips_dup += len(trips) - len(new_trips)

        for t in new_trips:
            code = t["trip_code"]
            db.add(models.OperationTrip(
                operation_id = op.id,
                trip_code    = code,
                entry_date   = _dt(t.get("entry_date")),
                entry_time   = t.get("entry_time"),
                exit_date    = _dt(t.get("exit_date")),
                exit_time    = t.get("exit_time"),
                plate        = t.get("plate"),
                tara_kg      = t.get("tara_kg"),
                bruto_kg     = t.get("bruto_kg"),
                neto_kg      = t.get("neto_kg"),
                origen_kg    = t.get("origen_kg"),
                diff_kg      = t.get("diff_kg"),
                shift_number = t.get("shift_number"),
                duration_min = t.get("duration_min"),
                client       = t.get("client"),
                product      = t.get("product"),
            ))
            existing_codes.add(code)  # evitar duplicados dentro del mismo archivo
            trips_inserted += 1

    db.commit()
    return {
        "ops_created":    ops_created,
        "trips_inserted": trips_inserted,
        "trips_dup":      trips_dup,
    }
