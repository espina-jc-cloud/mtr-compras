#!/usr/bin/env python3
"""
Crea tablas y genera el primer superadmin si no existe.
Se ejecuta automáticamente al iniciar en Railway.

Variables de entorno:
  FIRST_ADMIN_EMAIL     (default: admin@mtr.com)
  FIRST_ADMIN_PASSWORD  (requerida en producción — falla si no está)
"""
import os
import sys
from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import text
from app.database import engine, Base, DATABASE_URL, SessionLocal
from app import models
from app import models_cupos        # noqa: F401 — registra tablas de Despachos en Base.metadata
from app import models_tariffs      # noqa: F401 — registra tablas de Tarifario en Base.metadata
from app import models_transporte   # noqa: F401 — registra tablas de Transporte en Base.metadata
from app.auth import hash_password


def _sqlite_make_columns_nullable(conn, table, nullable_cols):
    """Reconstruye una tabla SQLite para que ciertas columnas pasen de NOT NULL a NULL.

    SQLite no soporta ALTER TABLE ... MODIFY COLUMN, así que el único camino
    es: crear tabla nueva con la definición correcta, copiar datos, borrar la
    vieja y renombrar.

    Solo se ejecuta si la DB es SQLite Y alguna columna todavía tiene notnull=1
    cuando debería ser nullable.  En PostgreSQL se omite completamente (las tablas
    se crean con la definición correcta desde el modelo desde el principio).

    Parámetros:
        conn           — conexión SQLAlchemy (raw) con autocommit desactivado
        table          — nombre de la tabla a reconstruir
        nullable_cols  — set/lista de columnas que deben quedar sin NOT NULL
    """
    # Obtener definición actual de columnas
    rows = conn.execute(text(f"PRAGMA table_info({table})")).fetchall()
    if not rows:
        return  # tabla no existe aún

    # Verificar si alguna de las columnas target todavía tiene notnull=1
    needs_fix = any(
        row[1] in nullable_cols and row[3] == 1   # row[1]=name, row[3]=notnull
        for row in rows
    )
    if not needs_fix:
        return  # nada que hacer

    print(f"  ~ reconstruyendo {table} para permitir NULL en {nullable_cols}...")

    # Construir CREATE TABLE para la tabla temporal con la definición corregida
    col_defs = []
    for row in rows:
        cid, name, col_type, notnull, dflt_value, pk = row
        pk_clause   = " PRIMARY KEY" if pk else ""
        dflt_clause = f" DEFAULT {dflt_value}" if dflt_value is not None else ""
        # Si la columna está en nullable_cols, forzar nullable (quitar NOT NULL)
        notnull_clause = "" if (name in nullable_cols or not notnull) else " NOT NULL"
        col_defs.append(f"  {name} {col_type}{pk_clause}{notnull_clause}{dflt_clause}")

    tmp = f"{table}__tmp_nullable"
    col_names = ", ".join(row[1] for row in rows)
    create_sql = f"CREATE TABLE {tmp} (\n" + ",\n".join(col_defs) + "\n)"

    conn.execute(text(f"DROP TABLE IF EXISTS {tmp}"))
    conn.execute(text(create_sql))
    conn.execute(text(f"INSERT INTO {tmp} ({col_names}) SELECT {col_names} FROM {table}"))
    conn.execute(text(f"DROP TABLE {table}"))
    conn.execute(text(f"ALTER TABLE {tmp} RENAME TO {table}"))
    conn.commit()
    print(f"  + {table} reconstruida: {nullable_cols} ahora aceptan NULL")


def _add_column(conn, table, column, col_type):
    """Agrega una columna si no existe. Compatible con SQLite y PostgreSQL.

    IMPORTANTE: el conn.rollback() en el except es obligatorio para PostgreSQL.
    Cuando ADD COLUMN falla (columna ya existe), PostgreSQL pone la conexión en
    estado "aborted transaction" — sin rollback, todos los comandos siguientes
    también fallan, aunque la columna no exista todavía.
    SQLite ignora el rollback sin error, así que es seguro para ambos motores.
    """
    try:
        conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}"))
        conn.commit()
        print(f"  + columna {table}.{column} agregada")
    except Exception:
        conn.rollback()  # ← crítico para PostgreSQL: limpia el estado "aborted"


def run():
    is_prod = not DATABASE_URL.startswith("sqlite")

    Base.metadata.create_all(bind=engine)
    # Las tablas `operations` y `operation_trips` se crean automáticamente aquí.
    print(f"✓ Tablas creadas ({DATABASE_URL.split('@')[-1] if '@' in DATABASE_URL else DATABASE_URL})")

    # ── Migraciones seguras de columnas nuevas ────────────────────────────────
    with engine.connect() as conn:
        _add_column(conn, "purchases",  "purchase_date",  "TIMESTAMP")
        _add_column(conn, "purchases",  "deleted_at",     "TIMESTAMP")
        _add_column(conn, "purchases",  "deleted_reason", "TEXT")
        _add_column(conn, "documents",  "remito_date",    "VARCHAR")
        _add_column(conn, "fuel_loads", "plant",          "VARCHAR")
        # Fase 2: Operativos en Tiempo Real — Cierre + Factura + Conciliación
        _add_column(conn, "operation_live_sessions", "closed_at",     "TIMESTAMP")
        _add_column(conn, "operation_live_sessions", "reconciled_at", "TIMESTAMP")
        _add_column(conn, "operation_live_shifts",       "turno_tipo",    "VARCHAR DEFAULT 'habil'")
        # Bodega: tipo de guinche y grampa por fila
        _add_column(conn, "operation_live_bodega_data", "tipo_guinche",  "VARCHAR")
        _add_column(conn, "operation_live_bodega_data", "tipo_grampa",   "VARCHAR")
        # Fase 3: Fotos — ya creada por Base.metadata.create_all() arriba;
        # _add_column aquí solo por si hay DBs viejas sin las columnas opcionales.
        # (La tabla en sí la crea create_all automáticamente.)
        # Tarifario v2: line_type / price_tier / visibility / adicionales / flags
        _add_column(conn, "tariffs", "line_type",           "VARCHAR(20) DEFAULT 'servicio'")
        _add_column(conn, "tariffs", "price_tier",          "VARCHAR(20) DEFAULT 'unica'")
        _add_column(conn, "tariffs", "visibility",          "VARCHAR(20) DEFAULT 'comercial'")
        _add_column(conn, "tariffs", "parent_id",           "INTEGER")
        _add_column(conn, "tariffs", "incluye_operador",    "BOOLEAN")
        _add_column(conn, "tariffs", "incluye_combustible", "BOOLEAN")
        _add_column(conn, "tariffs", "recargo_pct",         "NUMERIC(6,2)")
        _add_column(conn, "tariffs", "plaza",               "VARCHAR(120)")
        # Módulo Proyectos — Etapa 2A: Bitácora diaria
        # project_entries se crea sola con create_all().
        # Las columnas de projects que pasan a nullable no requieren ALTER en SQLite.
        # En PostgreSQL prod tampoco, porque la restricción en modelos es suficiente
        # (el constraint NOT NULL en columnas existentes no se altera automáticamente,
        # pero los registros existentes ya tienen valores y los nuevos son opcionales).
        # Módulo Proyectos — Etapa 2B: Adjuntos por entrada diaria
        # project_entry_attachments se crea sola con create_all().
        _add_column(conn, "project_entry_attachments", "currency", "VARCHAR")
        # SQLite: file_url y public_id fueron creadas NOT NULL antes de volverse nullable.
        # En PostgreSQL/Railway la tabla se crea directamente con la definición correcta.
        if not is_prod:
            _sqlite_make_columns_nullable(
                conn,
                "project_entry_attachments",
                {"file_url", "public_id"},
            )
        # Módulo Proyectos — Etapa Tareas
        # project_tasks se crea sola con create_all().
    print("✓ Columnas nuevas verificadas")

    admin_email = os.getenv("FIRST_ADMIN_EMAIL", "admin@mtr.com")
    admin_password = os.getenv("FIRST_ADMIN_PASSWORD", "")

    if not admin_password:
        if is_prod:
            print("✗ ERROR: FIRST_ADMIN_PASSWORD no está definida. Abortando.", file=sys.stderr)
            sys.exit(1)
        else:
            admin_password = "changeme123"
            print("⚠ Usando contraseña local por defecto: changeme123")

    # ── Seed catálogo de servicios del Tarifario (solo si está vacío) ──────────
    db_seed = SessionLocal()
    if db_seed.query(models_tariffs.TariffService).count() == 0:
        _servicios_seed = [
            ("Desestiba de buques",        "Movimiento de mercadería", "ton",   10),
            ("Carga de camiones",          "Movimiento de mercadería", "camion", 20),
            ("Descarga de camiones",       "Movimiento de mercadería", "camion", 30),
            ("Pesaje",                     "Servicios varios",          "camion", 40),
            ("Estiba en depósito",         "Depósito",                  "ton",   50),
            ("Almacenaje",                 "Depósito",                  "ton",   60),
            ("Uso exclusivo de depósito",  "Depósito",                  "mes",   70),
            ("Transporte",                 "Transporte",                "viaje", 80),
            ("Consolidado",                "Movimiento de mercadería", "contenedor", 90),
            ("Desconsolidado",             "Movimiento de mercadería", "contenedor", 100),
            ("Alquiler de equipos",        "Equipos",                   "dia",   110),
        ]
        for nombre, cat, unidad, orden in _servicios_seed:
            db_seed.add(models_tariffs.TariffService(
                nombre=nombre, categoria=cat, unidad_default=unidad, orden=orden, activo=True
            ))
        db_seed.commit()
        print(f"✓ Tarifario: {len(_servicios_seed)} servicios sembrados")
    else:
        print("✓ Tarifario: catálogo de servicios ya existe")
    db_seed.close()

    db = SessionLocal()
    existing = db.query(models.User).filter(models.User.email == admin_email).first()
    if not existing:
        admin = models.User(
            name="Administrador",
            email=admin_email,
            hashed_password=hash_password(admin_password),
            role="superadmin",
            plant="TODAS",
        )
        db.add(admin)
        db.commit()
        print(f"✓ Superadmin creado: {admin_email}")
    else:
        print(f"✓ Superadmin ya existe: {admin_email}")

    # Stats de tablas operativas
    counts = {
        "operations":               db.query(models.Operation).count(),
        "operation_trips":          db.query(models.OperationTrip).count(),
        "operation_product_totals": db.query(models.OperationProductTotal).count(),
        "operation_cargo_summaries": db.query(models.OperationCargoSummary).count(),
    }
    print(f"✓ Operativos: {counts['operations']} operations, {counts['operation_trips']} trips")
    print(f"  {counts['operation_product_totals']} product_totals (legacy) · "
          f"{counts['operation_cargo_summaries']} cargo_summaries (new)")

    db.close()

    # ── AUTO-IMPORT histórico combustible ─────────────────────────────────────
    # TEMPORAL: se eliminará después del primer deploy exitoso en producción.
    # Corre SOLO si:  prod + fuel_history.xlsx presente + tabla fuel_loads vacía.
    xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fuel_history.xlsx")
    if is_prod and os.path.exists(xlsx_path):
        _import_fuel_history(xlsx_path)

    # ── AUTO-IMPORT histórico operativos ──────────────────────────────────────
    # TEMPORAL: se eliminará después del primer deploy exitoso en producción.
    # Corre SOLO si:  prod + operations_history.xlsx presente + tablas vacías.
    ops_xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)), "operations_history.xlsx")
    if is_prod and os.path.exists(ops_xlsx):
        _import_operations_history(ops_xlsx)

    # ── AUTO-IMPORT costado vapor ─────────────────────────────────────────────
    # TEMPORAL: se eliminará después del primer deploy exitoso en producción.
    cv_xlsx = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cv_history.xlsx")
    if is_prod and os.path.exists(cv_xlsx):
        _import_cv_history(cv_xlsx)


def _import_fuel_history(xlsx_path: str):
    """Importa el historial de combustible desde el Excel exportado del Google Sheet."""
    import re
    import calendar
    from decimal import Decimal, InvalidOperation
    from datetime import datetime, timedelta

    try:
        import openpyxl
    except ImportError:
        print("✗ openpyxl no instalado — import combustible omitido")
        return

    db = SessionLocal()
    try:
        fuel_count = db.query(models.FuelLoad).count()
        if fuel_count > 0:
            print(f"✓ fuel_loads ya tiene {fuel_count} registros — import omitido")
            return

        # Usuario superadmin para entered_by
        sys_user = db.query(models.User).filter(models.User.role == "superadmin").first() \
                   or db.query(models.User).first()
        if not sys_user:
            print("✗ No hay usuarios — import combustible omitido")
            return

        # Helpers internos
        def _parse_decimal(val):
            if val is None:
                return None
            s = str(val).strip().replace(",", ".")
            if not s or s == "-":
                return None
            try:
                d = Decimal(s)
                return d if d > 0 else None
            except InvalidOperation:
                return None

        def _parse_date(val):
            if isinstance(val, datetime):
                return val.replace(hour=0, minute=0, second=0, microsecond=0)
            if val is None:
                return None
            s = str(val).strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
            return None

        FUEL_TYPE_MAP = {
            "gas oil premium": "gasoil_premium",
            "gasoil premium":  "gasoil_premium",
            "gasoil":          "gasoil_premium",
            "gas oil":         "gasoil_premium",
            "nafta premium":   "nafta_premium",
            "nafta":           "nafta",
            "super":           "nafta",
        }
        COMPANY_MAP = {
            "mtr sa": "MTR SA",
            "mtr":    "MTR SA",
            "ingee":  "INGEE",
            "ingée":  "INGEE",
        }

        def _map_fuel(val):
            k = str(val or "").strip().lower()
            for p, c in FUEL_TYPE_MAP.items():
                if p in k:
                    return c
            return "gasoil_premium"

        def _map_company(val):
            k = str(val or "").strip().lower()
            for p, n in COMPANY_MAP.items():
                if p in k:
                    return n
            return "MTR SA"

        def _infer_vtype(plate):
            p = plate.strip().upper()
            if p in ("BIDÓN", "BIDON", "BIDO", "BIDON 200L"):
                return "bidon"
            if re.match(r'^[A-Z]{2,3}\d{3}[A-Z]{0,2}$', p):
                return "vehiculo"
            return "equipo"

        def _order_num(val):
            if val is None:
                return None
            s = str(val).strip()
            if not s or s == "0":
                return None
            try:
                n = int(float(s))
                return str(n) if n > 0 else None
            except (ValueError, OverflowError):
                return s or None

        # Leer Excel
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = rows[0]

        def get(row, fragment):
            for j, h in enumerate(header):
                if h and fragment.lower() in str(h).lower():
                    return row[j] if j < len(row) else None
            return None

        inserted = skipped_bad = skipped_empty = 0

        for i, row in enumerate(rows[1:], 2):
            if all(v is None or str(v).strip() == "" for v in row):
                skipped_empty += 1
                continue

            plate = str(get(row, "patente") or "").strip().upper()
            fuel_dt = _parse_date(get(row, "fecha de carga") or get(row, "fecha"))
            liters = _parse_decimal(get(row, "litros"))

            if not plate or not fuel_dt or not liters:
                skipped_bad += 1
                print(f"  ⚠  fila {i}: incompleto (plate={plate!r} date={fuel_dt} liters={liters}) — omitida")
                continue
            if liters > 5000:
                skipped_bad += 1
                print(f"  ⚠  fila {i}: litros={liters} > 5000 (error de carga) — omitida")
                continue

            raw_ts = get(row, "marca temporal") or get(row, "timestamp")
            created = _parse_date(raw_ts) or fuel_dt

            load = models.FuelLoad(
                fuel_date        = fuel_dt,
                responsible_text = str(get(row, "responsable") or get(row, "nombre") or "").strip() or "Importado",
                entered_by_id    = sys_user.id,
                vehicle_plate    = plate,
                vehicle_type     = _infer_vtype(plate),
                fuel_type        = _map_fuel(get(row, "tipo de combustible") or get(row, "tipo")),
                liters           = liters,
                station          = str(get(row, "estaci") or "Hipolito").strip() or "Hipolito",
                amount           = _parse_decimal(get(row, "monto")),
                company          = _map_company(get(row, "empresa")),
                order_number     = _order_num(get(row, "orden")),
                created_at       = created,
                updated_at       = created,
            )
            db.add(load)
            inserted += 1

        db.commit()
        print(
            f"✓ Combustible importado: {inserted} registros · "
            f"{skipped_bad} omitidos (datos inválidos) · "
            f"{skipped_empty} filas vacías ignoradas"
        )

    except Exception as e:
        db.rollback()
        print(f"✗ Error en import combustible: {e}")
    finally:
        db.close()


def _import_operations_history(xlsx_path: str):
    """
    Importa el historial de operativos portuarios desde operations_history.xlsx.

    TEMPORAL — se elimina después del primer deploy exitoso en producción.
    Guard: prod + archivo presente + operations vacío + operation_trips vacío.

    Estructura del Excel (0-indexed):
      col[0]  = Nombre del Barco (header de operativo) / null en filas de viaje
      col[1]  = Cantidad de viajes (declarado, en header)
      col[2]  = código (int = viaje real; str = subtotal → ignorar)
      col[3]  = Fecha E (entrada)
      col[4]  = H Ent (hora entrada, time object)
      col[5]  = Fecha Sal (salida)
      col[6]  = H Sal (hora salida, time object)
      col[7]  = Patente camión
      col[8]  = Tara (kg)
      col[9]  = Bruto (kg)
      col[10] = Neto (kg)
      col[11] = Origen (kg)
      col[12] = Diferencias (kg)
      col[13] = Cliente
      col[14] = Producto
    """
    from datetime import datetime, timedelta
    from collections import Counter

    try:
        import openpyxl
    except ImportError:
        print("✗ openpyxl no instalado — import operativos omitido")
        return

    db = SessionLocal()
    try:
        ops_count   = db.query(models.Operation).count()
        trips_count = db.query(models.OperationTrip).count()
        if ops_count > 0 or trips_count > 0:
            print(f"✓ operations ya tiene {ops_count} operativos, {trips_count} viajes — import omitido")
            return

        # ── Normalización de nombres ──────────────────────────────────────────
        SHIP_ALIASES = {
            "MV ARGENMAR MISTRAL":  "ARGENMAR MISTRAL",
            "M/V ARGENMAR MISTRAL": "ARGENMAR MISTRAL",
        }
        SPECIAL_NAMES = {"INGRESO SOP", "ZONA FRANCA", "DEVOLUCIÓN BUNGE"}
        PRODUCT_FIX   = {"UREA GRANULA": "UREA GRANULADA"}

        def normalize_ship(name):
            s = name.strip()
            return SHIP_ALIASES.get(s, s)

        def fix_product(val):
            if val is None:
                return None
            s = str(val).strip()
            return PRODUCT_FIX.get(s, s) if s else None

        def get_shift(t_str):
            if not t_str:
                return None
            try:
                h = int(str(t_str)[:2])
                if h < 6:  return 1
                if h < 12: return 2
                if h < 18: return 3
                return 4
            except Exception:
                return None

        def calc_duration(entry_date, entry_time_str, exit_date, exit_time_str):
            if not all([entry_date, entry_time_str, exit_date, exit_time_str]):
                return None
            try:
                et = str(entry_time_str)[:8]  # HH:MM:SS
                xt = str(exit_time_str)[:8]
                entry_dt = datetime.combine(
                    entry_date if isinstance(entry_date, type(entry_date)) else entry_date,
                    datetime.strptime(et, "%H:%M:%S").time()
                )
                exit_dt = datetime.combine(
                    exit_date if isinstance(exit_date, type(exit_date)) else exit_date,
                    datetime.strptime(xt, "%H:%M:%S").time()
                )
                delta = exit_dt - entry_dt
                if delta.total_seconds() < 0:
                    delta += timedelta(days=1)
                mins = round(delta.total_seconds() / 60, 2)
                return mins if 0 <= mins <= 1440 else None
            except Exception:
                return None

        def most_common(lst):
            clean = [x for x in lst if x]
            return Counter(clean).most_common(1)[0][0] if clean else None

        def to_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.replace(hour=0, minute=0, second=0, microsecond=0)
            try:
                return datetime.strptime(str(val)[:10], "%Y-%m-%d")
            except Exception:
                return None

        def to_int(val):
            try:
                return int(val) if val is not None else None
            except Exception:
                return None

        # ── Leer Excel ────────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        wb.close()

        # ── Parse: agrupar filas por operativo ────────────────────────────────
        operations_data = []
        current_op      = None
        warnings        = []

        for row_idx, row in enumerate(raw_rows):
            def col(i):
                return row[i] if i < len(row) else None

            c0, c1, c2 = col(0), col(1), col(2)

            # Header de operativo: col[0] tiene nombre, col[2] es None o no numérico
            if c0 is not None and not isinstance(c2, (int, float)):
                raw_name = str(c0).strip()
                if raw_name:
                    current_op = {
                        "raw_name":      raw_name,
                        "ship_name":     normalize_ship(raw_name),
                        "operation_type":"special" if raw_name in SPECIAL_NAMES else "vessel",
                        "declared_trips": to_int(c1),
                        "trips":         [],
                    }
                    operations_data.append(current_op)
                continue

            # Fila de viaje real: col[2] es entero numérico
            if isinstance(c2, (int, float)) and c2 == int(c2):
                if current_op is None:
                    warnings.append(f"fila {row_idx+1}: viaje sin operativo padre — omitida")
                    continue

                entry_time_str = str(col(4))[:8] if col(4) is not None else None
                exit_time_str  = str(col(6))[:8] if col(6) is not None else None
                entry_date     = to_date(col(3))
                exit_date      = to_date(col(5))

                current_op["trips"].append({
                    "trip_code":    int(c2),
                    "entry_date":   entry_date,
                    "entry_time":   entry_time_str,
                    "exit_date":    exit_date,
                    "exit_time":    exit_time_str,
                    "plate":        str(col(7)).strip() if col(7) else None,
                    "tara_kg":      to_int(col(8)),
                    "bruto_kg":     to_int(col(9)),
                    "neto_kg":      to_int(col(10)),
                    "origen_kg":    to_int(col(11)),
                    "diff_kg":      to_int(col(12)),
                    "shift_number": get_shift(entry_time_str),
                    "duration_min": calc_duration(entry_date, entry_time_str, exit_date, exit_time_str),
                    "client":       str(col(13)).strip() if col(13) else None,
                    "product":      fix_product(col(14)),
                })
                continue
            # else: subtotal row → skip silently

        # ── Insert ────────────────────────────────────────────────────────────
        ops_inserted   = 0
        trips_inserted = 0
        trips_dup      = 0
        vessel_count   = 0
        special_count  = 0

        for op_data in operations_data:
            trips = op_data["trips"]
            if not trips:
                continue

            neto_list   = [t["neto_kg"]   for t in trips if t["neto_kg"]   is not None]
            origen_list = [t["origen_kg"] for t in trips if t["origen_kg"] is not None]
            diff_list   = [t["diff_kg"]   for t in trips if t["diff_kg"]   is not None]
            dur_list    = [t["duration_min"] for t in trips
                           if t["duration_min"] is not None and 0 < t["duration_min"] <= 240]

            total_neto    = sum(neto_list)
            actual_trips  = len(trips)
            avg_dur       = round(sum(dur_list)/len(dur_list), 2) if dur_list else None
            avg_t_trip    = round(total_neto/1000/actual_trips, 3) if actual_trips > 0 else None

            entry_dates = [t["entry_date"] for t in trips if t["entry_date"]]
            start_date  = min(entry_dates) if entry_dates else None
            end_date    = max(entry_dates) if entry_dates else None
            op_hours    = max((end_date-start_date).total_seconds()/3600, 1.0) if (start_date and end_date) else None
            avg_t_hour  = round(total_neto/1000/op_hours, 3) if op_hours else None

            op = models.Operation(
                raw_name          = op_data["raw_name"],
                ship_name         = op_data["ship_name"],
                operation_type    = op_data["operation_type"],
                client            = most_common([t["client"]   for t in trips]),
                product           = most_common([t["product"]  for t in trips]),
                start_date        = start_date,
                end_date          = end_date,
                declared_trips    = op_data["declared_trips"],
                actual_trips      = actual_trips,
                total_neto_kg     = total_neto,
                total_origen_kg   = sum(origen_list),
                total_diff_kg     = sum(diff_list),
                avg_duration_min  = avg_dur,
                avg_tons_per_trip = avg_t_trip,
                avg_tons_per_hour = avg_t_hour,
                source_file       = "operations_history.xlsx",
            )
            db.add(op)
            db.flush()
            ops_inserted += 1
            if op_data["operation_type"] == "vessel":
                vessel_count  += 1
            else:
                special_count += 1

            for t in trips:
                # Duplicate guard (trip_code UNIQUE — failsafe)
                existing = db.query(models.OperationTrip).filter(
                    models.OperationTrip.trip_code == t["trip_code"]
                ).first()
                if existing:
                    trips_dup += 1
                    continue
                db.add(models.OperationTrip(
                    operation_id  = op.id,
                    trip_code     = t["trip_code"],
                    entry_date    = t["entry_date"],
                    entry_time    = t["entry_time"],
                    exit_date     = t["exit_date"],
                    exit_time     = t["exit_time"],
                    plate         = t["plate"],
                    tara_kg       = t["tara_kg"],
                    bruto_kg      = t["bruto_kg"],
                    neto_kg       = t["neto_kg"],
                    origen_kg     = t["origen_kg"],
                    diff_kg       = t["diff_kg"],
                    shift_number  = t["shift_number"],
                    duration_min  = t["duration_min"],
                    client        = t["client"],
                    product       = t["product"],
                ))
                trips_inserted += 1

        db.commit()

        total_neto_t = sum(
            t["neto_kg"] for op in operations_data for t in op["trips"] if t["neto_kg"]
        ) / 1000

        print(f"✓ Operativos importados: {ops_inserted} operativos · {trips_inserted} viajes")
        print(f"  vessel={vessel_count} · special={special_count} · "
              f"total={total_neto_t:.1f} t · duplicados={trips_dup}")
        if warnings:
            print(f"  ⚠ {len(warnings)} avisos:")
            for w in warnings[:5]:
                print(f"    {w}")
            if len(warnings) > 5:
                print(f"    ... y {len(warnings)-5} más")

    except Exception as e:
        db.rollback()
        import traceback
        print(f"✗ Error en import operativos: {e}")
        traceback.print_exc()
    finally:
        db.close()


def _import_cv_history(xlsx_path: str):
    """
    Importa el historial de Costado Vapor desde cv_history.xlsx.

    TEMPORAL — se elimina después del primer deploy exitoso en producción.
    Guard: operation_product_totals vacío.
    """
    import unicodedata
    from collections import defaultdict
    from datetime import datetime

    try:
        import openpyxl
    except ImportError:
        print("✗ openpyxl no instalado — import CV omitido")
        return

    db = SessionLocal()
    try:
        cv_count = db.query(models.OperationProductTotal).count()
        if cv_count > 0:
            print(f"✓ operation_product_totals ya tiene {cv_count} registros — import CV omitido")
            return

        SOURCE_FILE = "cv_history.xlsx"
        DATE_WINDOW = 10

        PRODUCT_FIX_CV = {
            "Súoer Fosfato Triple":   "TRIPLE (STP)",
            "Súper Fosfato Triple":   "TRIPLE (STP)",
            "Super Fosfato Triple":   "TRIPLE (STP)",
            "Fosfato Monoamònico":    "MAP",
            "Fosfato Monoamónico":    "MAP",
            "Fosfato Diamónico":      "DAP",
            "Urea Granulada":         "UREA GRANULADA",
            "Cloturo de Potasio":     "CLORURO DE POTASIO",
            "Sulfato de Amonio":      "SULFATO DE AMONIO",
        }

        DEPOT_PRODUCT_ALIASES = {
            "MOP":   "CLORURO DE POTASIO",
            "AMSUL": "SULFATO DE AMONIO",
        }

        SHIP_FIX = {
            "M/V Sider Liu":        "SIDER LIU",
            "M/V Arriabbiata":      "ARRABIATA",
            "M/V Ruen":             "MV/RUEN",
            "M/V Argenmar Mistral": "ARGENMAR MISTRAL",
            "M/V Seacon Bangkok":   "SEACON BANGKOK",
            "M/V Nava Ulysses":     "NAVA ULYSSES",
            "M/V Alithia":          "M/V ALITHIA",
            "M/V Genius Star IX":   "GENIUS STAR IX",
            "M/V Amani":            "M/V AMANI",
            "M/V Areti":            "M/V ARETI",
            "M/V CL ZHANGJIAJIE":  "M/V CL ZHANGJIAJIE",
            "M/V Endless Horizon":  "M/V ENDLESS HORIZON",
            "M/V Yasa Moon\xa0":   "YASA MOON",
            "M/V Beatrice":         "M/V BEATRICE",
            "M/V Bonny Island":     "M/V BONNY ISLAND",
            "M/V Sofía":            "M/V SOFIA",
            "M/V Vega":             "M/V VEGA",
            "M/V Ultralaz":         "M/V ULTRALAZ",
            "M/V Tenca Arrow":      "M/V TENCA ARROW",
            "M/V Adasun":           "ADASUN",
            "M/V IC Progress":      "M/V IC PROGRESS",
            "M/V Oborishte":        "OBORISHTE",
            "M/V ATHANASIA":        "ATHANASIA",
            "M/V NIKITIS":          "NIKITIS",
        }

        def norm_key(v):
            if not v:
                return ""
            s = unicodedata.normalize("NFD", str(v).strip())
            s = "".join(c for c in s if unicodedata.category(c) != "Mn")
            s = s.upper()
            for pre in ("M/V ", "MV/", "MV "):
                if s.startswith(pre):
                    s = s[len(pre):]
            return s.strip()

        def fix_product_cv(raw):
            if raw is None:
                return "(sin producto)"
            s = str(raw).strip()
            return PRODUCT_FIX_CV.get(s, s.upper() if s else "(sin producto)")

        def to_dt(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.replace(hour=0, minute=0, second=0, microsecond=0)
            try:
                return datetime.strptime(str(val)[:10], "%Y-%m-%d")
            except Exception:
                return None

        # ── Parse Excel ───────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        data_rows = rows[7:37]  # 0-indexed 7 to 36 inclusive

        records = []
        last_ship = None
        last_start = None
        last_end = None

        for i, row in enumerate(data_rows):
            def col(j, r=row):
                return r[j] if j < len(r) else None

            raw_ship   = col(1)
            raw_start  = col(2)
            raw_end    = col(3)
            raw_client = col(5)
            raw_product = col(6)
            raw_tons   = col(7)

            if raw_ship and str(raw_ship).strip():
                last_ship  = str(raw_ship).strip()
                last_start = raw_start
                last_end   = raw_end
            else:
                raw_ship  = last_ship
                raw_start = last_start
                raw_end   = last_end

            if not raw_ship:
                continue
            if raw_tons is None:
                continue
            try:
                tons = float(raw_tons)
            except (TypeError, ValueError):
                continue
            if tons <= 0:
                continue

            start_dt = to_dt(raw_start)
            end_dt   = to_dt(raw_end)
            ship_norm = SHIP_FIX.get(raw_ship, raw_ship)

            if "AMANI" in raw_ship.upper() or "AMANI" in ship_norm.upper():
                if start_dt and start_dt.year == 2026:
                    start_dt = start_dt.replace(year=2025)
                if end_dt and end_dt.year == 2026:
                    end_dt = end_dt.replace(year=2025)

            if "NIKITIS" in raw_ship.upper() or "NIKITIS" in ship_norm.upper():
                if start_dt and end_dt and end_dt < start_dt:
                    start_dt, end_dt = end_dt, start_dt

            product_norm = fix_product_cv(raw_product)
            client_str   = str(raw_client).strip() if raw_client else None

            records.append({
                "raw_ship_name": raw_ship,
                "ship_name":     ship_norm,
                "match_key":     norm_key(ship_norm),
                "client":        client_str,
                "product":       product_norm,
                "cv_start_date": start_dt,
                "cv_end_date":   end_dt,
                "cv_excel_tons": tons,
            })

        # ── Load ops + depot tons ─────────────────────────────────────────────
        all_ops   = db.query(models.Operation).all()
        all_trips = db.query(models.OperationTrip).all()

        depot_by_op_prod = defaultdict(lambda: defaultdict(float))
        for t in all_trips:
            if t.neto_kg:
                pn = DEPOT_PRODUCT_ALIASES.get(t.product, t.product) if t.product else "(sin producto)"
                depot_by_op_prod[t.operation_id][pn] += t.neto_kg / 1000

        # ── Match + insert ────────────────────────────────────────────────────
        matched = unmatched = inserted = 0
        total_vapor = 0.0

        for rec in records:
            mk = rec["match_key"]
            start_dt = rec["cv_start_date"]
            prod = rec["product"]

            candidates = [op for op in all_ops if norm_key(op.ship_name) == mk]
            op = None
            status = "unmatched"
            notes = None

            if candidates:
                if start_dt:
                    scored = []
                    for c in candidates:
                        ref = c.start_date or c.end_date
                        dist = abs((start_dt - ref).days) if ref else DATE_WINDOW
                        if dist <= DATE_WINDOW:
                            has_prod = 0 if prod in depot_by_op_prod.get(c.id, {}) else 1
                            scored.append((has_prod, dist, c))
                    if not scored:
                        scored = [(1, 999, c) for c in candidates]
                else:
                    scored = [(1, 0, c) for c in candidates]

                scored.sort(key=lambda x: (x[0], x[1]))
                best_score, best_dist, best_op = scored[0]
                ties = [x for x in scored if x[0] == best_score and x[1] == best_dist]
                op = best_op
                if len(ties) > 1:
                    status = "ambiguous"
                    notes = f"Ambiguous: {len(ties)} ops"
                else:
                    status = "matched"
                matched += 1
            else:
                unmatched += 1
                notes = f"No op for '{rec['ship_name']}'"

            op_id   = op.id if op else None
            depot_t = depot_by_op_prod.get(op_id, {}).get(prod, 0.0) if op_id else 0.0
            cv_t    = rec["cv_excel_tons"]
            vapor_t = max(cv_t - depot_t, 0.0)
            total_t = cv_t if cv_t >= depot_t else depot_t
            total_vapor += vapor_t

            db.add(models.OperationProductTotal(
                operation_id          = op_id,
                raw_ship_name         = rec["raw_ship_name"],
                ship_name             = rec["ship_name"],
                client                = rec["client"],
                product               = prod,
                cv_start_date         = rec["cv_start_date"],
                cv_end_date           = rec["cv_end_date"],
                cv_excel_tons         = cv_t,
                depot_tons            = depot_t,
                costado_vapor_tons    = vapor_t,
                total_discharged_tons = total_t,
                match_status          = status,
                notes                 = notes,
                source_file           = SOURCE_FILE,
                source_year           = rec["cv_start_date"].year if rec["cv_start_date"] else None,
            ))
            inserted += 1

        db.commit()
        print(
            f"✓ CV importado: {inserted} registros · "
            f"matched={matched} · unmatched={unmatched} · "
            f"vapor={total_vapor:.0f} t"
        )

    except Exception as e:
        db.rollback()
        import traceback
        print(f"✗ Error en import CV: {e}")
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    run()
